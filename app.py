"""
标注评测报告生成工具 - Streamlit主界面
"""

import os
import copy
import re
import difflib
import streamlit as st
import pandas as pd
import io
from typing import Optional, Tuple, List, Dict, Any
from datetime import datetime

# 读取 .env 文件（如果存在）
_env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(_env_path):
    with open(_env_path, "r", encoding="utf-8") as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())
# 导入新架构模块
from services_2_1 import DataService
from domain_1 import MetricsConfig

def get_column_unique_values(df, column: str):
    return DataService(df).get_column_unique_values(column)
from services_2_2 import LLMService, format_api_error
from prompts_3 import (
    PARSE_METRIC_EXCEL_PROMPT,
    PARSE_METRIC_GSB_PROMPT,
    REPORT_PROMPT_TEMPLATE,
    EMPTY_METRICS_CONFIG
)
from exports_6 import ExportService
from utils_7 import format_stats_for_prompt
# from domain_1 import MetricsConfig, Metric, Condition  # 未使用，已移除
from utils_7 import format_stats_result_for_display, format_stats_result_for_grouping

# ==================== 页面配置 ====================
st.set_page_config(
    page_title="标注评测报告生成工具",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== 自定义CSS（极简现代风格） ====================
# 字体：Inter - 来自 Google Fonts
st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
    /* ==================== 设计系统变量 ==================== */
    :root {
        --primary: #0F172A;
        --primary-hover: #1E293B;
        --secondary: #64748B;
        --accent: #10B981;
        --accent-hover: #059669;
        --gray-50: #F9FAFB;
        --gray-100: #F3F4F6;
        --gray-200: #E5E7EB;
        --gray-300: #D1D5DB;
        --gray-400: #9CA3AF;
        --gray-500: #6B7280;
        --gray-600: #4B5563;
        --gray-700: #374151;
        --gray-800: #1F2937;
        --gray-900: #111827;
        --success: #10B981;
        --warning: #F59E0B;
        --error: #EF4444;
        --info: #3B82F6;
        --radius-sm: 4px;
        --radius-md: 8px;
        --radius-lg: 12px;
    }

    /* ==================== 全局基础 ==================== */
    /* 强制所有元素使用 Inter 字体 */
    *, *::before, *::after {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif !important;
    }
    .stApp {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif;
        background-color: var(--gray-50);
        color: var(--gray-900);
        -webkit-font-smoothing: antialiased;
        -moz-osx-font-smoothing: grayscale;
    }
    
    /* 核心容器排版 */
    .block-container {
        max-width: 1200px;
        padding-top: 2rem;
        padding-bottom: 4rem;
    }
    
    /* 文本与层级 */
    h1, h2, h3, h4, h5, h6 {
        color: var(--primary);
        font-weight: 600;
        letter-spacing: -0.01em;
    }
    h1 { font-size: 2.5rem; font-weight: 700; }
    h2 { font-size: 1.5rem; margin-top: 1.5rem; margin-bottom: 1rem; }
    h3 { font-size: 1.25rem; margin-top: 1.25rem; margin-bottom: 0.75rem; }

    /* 正文文字 */
    .stMarkdown p, li, label, .stCaption {
        color: var(--gray-600);
        line-height: 1.6;
        font-size: 14px;
    }

    /* 卡片与表单容器 */
    [data-testid="stForm"], [data-testid="stExpander"] {
        background: #FFFFFF;
        border: 1px solid var(--gray-200) !important;
        border-radius: var(--radius-lg) !important;
        box-shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.05) !important;
        padding: 1.25rem;
    }
    [data-testid="stExpander"] details summary {
        font-weight: 600;
        color: var(--gray-900);
    }

    /* ==================== 交互控件 ==================== */
    /* 输入框 */
    .stTextInput input, .stTextArea textarea, .stSelectbox [data-baseweb="select"] > div, .stMultiSelect [data-baseweb="select"] > div, .stNumberInput input {
        border: 1px solid var(--gray-300) !important;
        border-radius: var(--radius-md) !important;
        background: #FFFFFF !important;
        color: var(--gray-900) !important;
        font-size: 14px;
        transition: all 0.15s ease;
        box-shadow: 0 1px 2px 0 rgba(0, 0, 0, 0.02) !important;
    }
    .stTextInput input:focus, .stTextArea textarea:focus, .stNumberInput input:focus {
        border-color: var(--accent) !important;
        box-shadow: 0 0 0 3px rgba(16, 185, 129, 0.1) !important;
    }

    /* 按钮基础风格 */
    .stButton button {
        border-radius: var(--radius-md);
        border: 1px solid var(--gray-300);
        background-color: #FFFFFF;
        font-weight: 500;
        font-size: 14px;
        min-height: 2.5rem;
        padding: 0.5rem 1rem;
        box-shadow: 0 1px 2px 0 rgba(0, 0, 0, 0.03);
        transition: all 0.15s ease;
    }
    .stButton button p {
        color: inherit !important;
        margin: 0 !important;
        font-weight: inherit !important;
    }
    .stButton button:hover {
        border-color: var(--gray-400);
        background-color: var(--gray-50);
        color: var(--gray-900);
    }

    /* 主按钮（深灰主色） */
    .stButton button[data-testid="baseButton-primary"], .stButton button[kind="primary"] {
        background-color: var(--primary) !important;
        border-color: var(--primary) !important;
        color: #FFFFFF !important;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.08) !important;
    }
    .stButton button[data-testid="baseButton-primary"]:hover, .stButton button[kind="primary"]:hover {
        background-color: var(--primary-hover) !important;
        border-color: var(--primary-hover) !important;
        transform: translateY(-1px);
    }

    /* 强调按钮（翠绿） */
    .stButton button.accent-button {
        background-color: var(--accent) !important;
        border-color: var(--accent) !important;
        color: #FFFFFF !important;
    }
    .stButton button.accent-button:hover {
        background-color: var(--accent-hover) !important;
        transform: translateY(-1px);
    }

    /* 禁用态按钮 */
    .stButton button:disabled, .stButton button[disabled] {
        opacity: 0.6 !important;
        cursor: not-allowed !important;
        transform: none !important;
        box-shadow: none !important;
    }

    /* 侧边栏 */
    [data-testid="stSidebar"] {
        background-color: var(--gray-100);
        border-right: 1px solid var(--gray-200);
    }

    /* 标签页 */
    .stTabs button {
        font-size: 14px;
        font-weight: 500;
        color: var(--gray-500);
    }
    .stTabs button[aria-selected="true"] {
        color: var(--primary) !important;
        border-bottom-color: var(--primary) !important;
    }

    /* 分割线 */
    .stDivider {
        border-bottom-color: var(--gray-200);
        margin-top: 2rem;
        margin-bottom: 2rem;
    }

    /* 上传区优化 */
    .stFileUploader label[data-testid="stFileUploaderDropzone"] {
        background: #FFFFFF;
        border: 2px dashed var(--gray-300);
        border-radius: var(--radius-lg);
        transition: border-color 0.2s ease;
        padding: 2rem;
    }
    .stFileUploader label[data-testid="stFileUploaderDropzone"]:hover {
        border-color: var(--accent);
        background: rgba(16, 185, 129, 0.02);
    }

    /* 隐藏默认头部 */
    header[data-testid="stHeader"], #MainMenu {
        display: none !important;
    }

    /* ==================== 数据表格美化 ==================== */
    .stDataFrame {
        border: 1px solid var(--gray-200);
        border-radius: var(--radius-lg);
        overflow: hidden;
    }
    .stDataFrame thead th {
        background: var(--gray-50);
        font-weight: 600;
        color: var(--gray-700);
        border-bottom: 2px solid var(--gray-200);
    }
    .stDataFrame tbody tr:hover {
        background: var(--gray-50);
    }

    /* ==================== 徽章系统 ==================== */
    .badge {
        display: inline-flex;
        align-items: center;
        padding: 2px 8px;
        border-radius: 9999px;
        font-size: 12px;
        font-weight: 500;
    }
    .badge-success { background: #D1FAE5; color: #065F46; }
    .badge-warning { background: #FEF3C7; color: #92400E; }
    .badge-error { background: #FEE2E2; color: #991B1B; }
    .badge-info { background: #DBEAFE; color: #1E40AF; }

    /* ==================== 步骤指示器 ==================== */
    .steps-container {
        display: flex;
        align-items: center;
        background: #FFFFFF;
        border: 1px solid var(--gray-200);
        border-radius: var(--radius-lg);
        padding: 12px 16px;
        margin-bottom: 1.5rem;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
    }
    .step {
        display: flex;
        align-items: center;
        flex: 1;
    }
    .step-number {
        width: 28px;
        height: 28px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 12px;
        font-weight: 600;
        margin-right: 8px;
    }
    .step.completed .step-number { background: var(--accent); color: white; }
    .step.current .step-number { background: var(--primary); color: white; }
    .step.pending .step-number { background: var(--gray-200); color: var(--gray-500); }
    .step-text {
        font-size: 14px;
        color: var(--gray-600);
    }
    .step.completed .step-text, .step.current .step-text {
        color: var(--gray-900);
        font-weight: 500;
    }
    .step-line {
        flex: 1;
        height: 2px;
        background: var(--gray-200);
        margin: 0 12px;
    }
    .step.completed + .step-line { background: var(--accent); }
        display: inline-flex;
        align-items: center;
        padding: 0.375rem 0.75rem;
        border-radius: 9999px;
        font-size: 13px;
        font-weight: 500;
        background: #F3F4F6;
        color: #6B7280;
        border: 1px solid #E5E7EB;
    }

    .status-badge.success {
        background: #ECFDF5;
        color: #10B981;
        border-color: #D1FAE5;
    }

    .status-badge.active {
        background: #EFF6FF;
        color: #3B82F6;
        border-color: #BFDBFE;
    }

    /* ==================== 卡片式场景选择 ==================== */
    .scene-card {
        background: #FFFFFF;
        border: 2px solid #E5E7EB;
        border-radius: 12px;
        padding: 1.5rem;
        text-align: center;
        cursor: pointer;
        transition: all 0.2s ease;
        min-height: 140px;
    }

    .scene-card:hover {
        border-color: #9CA3AF;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08);
    }

    .scene-card.selected {
        border-color: #0F172A;
        background: #F9FAFB;
    }

    .scene-card .icon {
        font-size: 2.5rem;
        margin-bottom: 0.75rem;
    }

    .scene-card .title {
        font-size: 1rem;
        font-weight: 600;
        color: #111827;
        margin-bottom: 0.5rem;
    }

    .scene-card .desc {
        font-size: 13px;
        color: #6B7280;
        line-height: 1.4;
    }

    /* ==================== 空状态设计 ==================== */
    .empty-state {
        text-align: center;
        padding: 4rem 2rem;
        background: #FFFFFF;
        border: 1px solid #E5E7EB;
        border-radius: 12px;
        margin: 2rem 0;
    }

    .empty-icon {
        font-size: 4rem;
        margin-bottom: 1rem;
        opacity: 0.6;
    }

    .empty-title {
        font-size: 1.25rem;
        font-weight: 600;
        color: #111827;
        margin-bottom: 0.5rem;
    }

    .empty-desc {
        font-size: 14px;
        color: #6B7280;
        margin-bottom: 1.5rem;
    }

    /* ==================== 加载状态 ==================== */
    .loading-state {
        text-align: center;
        padding: 3rem 2rem;
        background: #FFFFFF;
        border: 1px solid #E5E7EB;
        border-radius: 12px;
        margin: 2rem 0;
    }

    .loading-text {
        font-size: 1rem;
        font-weight: 500;
        color: #111827;
        margin-top: 1rem;
    }

    .loading-hint {
        font-size: 13px;
        color: #6B7280;
        margin-top: 0.5rem;
    }

    /* ==================== 按钮分层系统 ==================== */
    /* 主按钮 - Primary */
    .stButton button[kind="primary"] {
        background: #0F172A !important;
        color: #FFFFFF !important;
        border: none !important;
        font-weight: 600;
        padding: 0.75rem 1.5rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1) !important;
    }

    .stButton button[kind="primary"]:hover {
        background: #1E293B !important;
        transform: translateY(-1px);
        box-shadow: 0 6px 8px -1px rgba(0, 0, 0, 0.12) !important;
    }

    /* 次按钮 - Secondary */
    .stButton button.secondary {
        background: #FFFFFF !important;
        color: #0F172A !important;
        border: 1px solid #D1D5DB !important;
        font-weight: 500;
    }

    .stButton button.secondary:hover {
        background: #F9FAFB !important;
        border-color: #9CA3AF !important;
    }

    /* 危险按钮 - Danger */
    .stButton button.danger {
        background: #FFFFFF !important;
        color: #EF4444 !important;
        border: 1px solid #FCA5A5 !important;
        font-weight: 500;
    }

    .stButton button.danger:hover {
        background: #FEF2F2 !important;
        border-color: #F87171 !important;
    }

    /* ==================== 实时统计预览 ==================== */
    .metrics-summary {
        background: #FFFFFF;
        border: 1px solid #E5E7EB;
        border-radius: 12px;
        padding: 1.25rem;
        margin: 1.5rem 0;
        box-shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.05);
    }

    .metrics-summary-title {
        font-size: 0.875rem;
        font-weight: 600;
        color: #6B7280;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 1rem;
    }

    /* ==================== 智能提示系统 ==================== */
    .smart-hint {
        background: #EFF6FF;
        border: 1px solid #BFDBFE;
        border-left: 4px solid #3B82F6;
        border-radius: 8px;
        padding: 1rem 1.25rem;
        margin: 1.5rem 0;
    }

    .smart-hint-text {
        font-size: 14px;
        color: #1E40AF;
        font-weight: 500;
    }

    /* ==================== 数据表格美化 ==================== */
    .stDataFrame {
        border: 1px solid #E5E7EB;
        border-radius: 8px;
        overflow: hidden;
    }

    /* ==================== 微交互 ==================== */
    .stButton button {
        transition: all 0.2s ease;
    }

    .stButton button:active {
        transform: translateY(0);
    }

    /* ==================== 语义化颜色变量 ==================== */
    :root {
        --color-success: #10B981;
        --color-warning: #F59E0B;
        --color-error: #EF4444;
        --color-info: #3B82F6;
        --color-primary: #0F172A;
        --color-secondary: #64748B;
        --color-neutral: #6B7280;
    }
</style>
""", unsafe_allow_html=True)

# ==================== 辅助函数 ====================

def load_api_config():
    """从 .api_config.json 加载已保存的API配置"""
    import json
    config_path = os.path.join(os.path.dirname(__file__), ".api_config.json")
    if os.path.exists(config_path):
        with open(config_path, "r") as f:
            return json.load(f)
    return {}

def save_api_config(api_key, base_url, model_name):
    """保存API配置到 .api_config.json"""
    import json
    config_path = os.path.join(os.path.dirname(__file__), ".api_config.json")
    with open(config_path, "w") as f:
        json.dump({"api_key": api_key, "base_url": base_url, "model_name": model_name}, f)

def delete_api_config():
    """删除已保存的API配置"""
    config_path = os.path.join(os.path.dirname(__file__), ".api_config.json")
    if os.path.exists(config_path):
        os.remove(config_path)



def remove_duplicates(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    """删除 DataFrame 中的重复行

    Args:
        df: 输入的 DataFrame

    Returns:
        Tuple[pd.DataFrame, int]: 去重后的 DataFrame 和删除的行数
    """
    original_count = len(df)
    df_deduped = df.drop_duplicates()
    removed_count = original_count - len(df_deduped)
    return df_deduped, removed_count


def get_column_summary(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """获取 DataFrame 各列的摘要信息

    Args:
        df: 输入的 DataFrame

    Returns:
        List[Dict]: 每列的摘要信息列表
    """
    summary = []
    for col in df.columns:
        series = df[col]
        non_null_count = series.notna().sum()
        unique_count = series.nunique()
        summary.append({
            "列名": col,
            "非空值数": int(non_null_count),
            "空值数": int(len(df) - non_null_count),
            "唯一值数": int(unique_count),
            "类型": str(series.dtype)
        })
    return summary


def init_session_state():
    """初始化session_state"""
    if "df" not in st.session_state:
        st.session_state["df"] = None
    if "columns" not in st.session_state:
        st.session_state["columns"] = []
    if "parsed_metrics" not in st.session_state:
        st.session_state["parsed_metrics"] = None
    if "confirmed_metrics" not in st.session_state:
        st.session_state["confirmed_metrics"] = None
    if "editing_metrics" not in st.session_state:
        st.session_state["editing_metrics"] = None
    if "stats_result" not in st.session_state:
        st.session_state["stats_result"] = None
    if "comparison_mode" not in st.session_state:
        st.session_state["comparison_mode"] = False
    if "generated_report" not in st.session_state:
        st.session_state["generated_report"] = ""
    if "is_generating_report" not in st.session_state:
        st.session_state["is_generating_report"] = False
    if "trigger_parse_excel" not in st.session_state:
        st.session_state["trigger_parse_excel"] = False
    # 加载已保存的API配置
    saved = load_api_config()
    if saved:
        st.session_state.setdefault("api_key", saved.get("api_key", ""))
        st.session_state.setdefault("base_url", saved.get("base_url", ""))
        st.session_state.setdefault("model_name", saved.get("model_name", ""))
    # Case筛选相关
    if "badcase_conditions" not in st.session_state:
        st.session_state["badcase_conditions"] = []
    if "badcase_conditions_prev" not in st.session_state:
        st.session_state["badcase_conditions_prev"] = []
    if "goodcase_conditions" not in st.session_state:
        st.session_state["goodcase_conditions"] = []
    if "goodcase_conditions_prev" not in st.session_state:
        st.session_state["goodcase_conditions_prev"] = []
    if "badcase_result" not in st.session_state:
        st.session_state["badcase_result"] = None
    if "goodcase_result" not in st.session_state:
        st.session_state["goodcase_result"] = None


def validate_session_keys():
    """验证session_state键是否正常"""
    pass  # 此函数暂时禁用，避免不必要的检查


def clear_generated_report():
    st.session_state["generated_report"] = ""
    st.session_state["is_generating_report"] = False


def request_parse_excel():
    st.session_state["trigger_parse_excel"] = True


def clear_metrics_editor_keys():
    prefixes = (
        "denom_cond_", "numer_cond_", "or_cond_", "metric_name_", "denom_type_",
        "ai_denom_cond_", "denom_grp_cond_", "metric_ai_denom_", "custom_denom_",
        "denom_source_", "denom_grp_sel_", "denom_grp_name_"
    )
    exact_keys = (
        "denom_desc", "denom_type_radio", "use_ai_denom_btn", "clear_denom_btn",
        "add_denom_grp", "add_denom_cond"
    )
    keys_to_remove = [k for k in st.session_state if k.startswith(prefixes) or k in exact_keys]
    for k in keys_to_remove:
        del st.session_state[k]


def get_eval_field_distribution(df):
    """只返回用户勾选的评估字段的分布信息"""
    import json
    full_dist = DataService(df).build_field_distribution()
    eval_fields = st.session_state.get("eval_fields", list(full_dist.keys()))
    filtered = {k: v for k, v in full_dist.items() if k in eval_fields}
    return json.dumps(filtered, ensure_ascii=False, indent=2)


def auto_fix_operators(parsed_config: dict, df) -> dict:
    """根据字段分布自动修正运算符（如 multi_select 字段的 in → contains）"""
    field_dist = DataService(df).build_field_distribution()

    def fix_condition(cond):
        field = cond.get("field", "")
        op = cond.get("op", "")
        info = field_dist.get(field, {})
        if info.get("type") == "categorical_with_multi":
            if op in ("in", "=="):
                cond["op"] = "contains"
            elif op in ("not_in", "!="):
                cond["op"] = "not_contains"

    # 修正公共分母
    for cond in parsed_config.get("common_denominator", {}).get("conditions", []):
        fix_condition(cond)

    # 修正每个指标
    for metric in parsed_config.get("metrics", []):
        for cond in metric.get("numerator_conditions", []):
            fix_condition(cond)
        for group in metric.get("numerator_or_conditions", []):
            for cond in group:
                fix_condition(cond)
        for cond in metric.get("custom_denominator_conditions", []):
            fix_condition(cond)

    return parsed_config


def auto_fix_values(parsed_config, df):
    """根据字段分布自动修正条件中的值（模糊匹配），返回 (修正后config, 修正日志列表)"""
    import difflib
    field_dist = DataService(df).build_field_distribution()
    fix_log = []

    def get_valid_values(info):
        t = info.get("type", "")
        if t == "categorical":
            return info.get("values", [])
        elif t == "categorical_with_multi":
            return info.get("options", [])
        return []

    def fix_value(cond):
        field = cond.get("field", "")
        op = cond.get("op", "")
        val = cond.get("value", "")
        if op in ("is_empty", "is_not_empty") or not val:
            return
        info = field_dist.get(field, {})
        valid = get_valid_values(info)
        if not valid:
            return

        if op in ("in", "not_in"):
            parts = [v.strip() for v in val.split(",")]
            new_parts = []
            for p in parts:
                if p in valid:
                    new_parts.append(p)
                else:
                    min_cutoff = 0.85 if len(p) <= 4 else 0.7
                    match = difflib.get_close_matches(p, valid, n=1, cutoff=min_cutoff)
                    if match:
                        ratio = difflib.SequenceMatcher(None, p, match[0]).ratio()
                        fix_log.append({"field": field, "old": p, "new": match[0], "ratio": round(ratio * 100)})
                        new_parts.append(match[0])
                    else:
                        new_parts.append(p)
            cond["value"] = ",".join(new_parts)
        else:
            if val not in valid:
                min_cutoff = 0.85 if len(val) <= 4 else 0.7
                match = difflib.get_close_matches(val, valid, n=1, cutoff=min_cutoff)
                if match:
                    ratio = difflib.SequenceMatcher(None, val, match[0]).ratio()
                    fix_log.append({"field": field, "old": val, "new": match[0], "ratio": round(ratio * 100)})
                    cond["value"] = match[0]

    for cond in parsed_config.get("common_denominator", {}).get("conditions", []):
        fix_value(cond)
    for metric in parsed_config.get("metrics", []):
        for cond in metric.get("numerator_conditions", []):
            fix_value(cond)
        for group in metric.get("numerator_or_conditions", []):
            for cond in group:
                fix_value(cond)
        for cond in metric.get("custom_denominator_conditions", []):
            fix_value(cond)

    return parsed_config, fix_log


def auto_fix_level_codes(parsed_config, df, spec_text):
    """修正条件值中混淆的L等级代码（如L3写成L0）

    逻辑：在 spec_text 中找到与条件关键词相关的行，
    提取该行出现的实际L等级，若与条件值里的等级不符则替换。
    返回 (fixed_config, fix_log)
    """
    import re
    field_dist = DataService(df).build_field_distribution()
    fix_log = []

    def get_valid_values(info):
        t = info.get("type", "")
        if t == "categorical":
            return info.get("values", [])
        elif t == "categorical_with_multi":
            return info.get("options", [])
        return []

    # 从 spec_text 中提取所有出现的 L\d 等级
    spec_lines = spec_text.split("\n") if spec_text else []

    def find_levels_in_spec(keyword):
        """在spec中找含keyword的行，提取所有L\d等级"""
        levels = set()
        for line in spec_lines:
            if keyword in line:
                found = re.findall(r'L\d', line)
                levels.update(found)
        return levels

    def fix_cond(cond):
        field = cond.get("field", "")
        op = cond.get("op", "")
        val = cond.get("value", "")
        if op in ("is_empty", "is_not_empty") or not val:
            return
        info = field_dist.get(field, {})
        valid = get_valid_values(info)
        if not valid:
            return

        # 找val中的L等级
        level_in_val = re.search(r'L(\d)', val)
        if not level_in_val:
            return

        # val已经合法，跳过
        if val in valid:
            return

        # 从val中提取中文关键词（去掉Lx部分，取剩余2字）
        keyword = re.sub(r'L\d', '', val).strip()[:2]
        if not keyword:
            return

        # 在spec里找该关键词相关行的实际等级
        spec_levels = find_levels_in_spec(keyword)
        if not spec_levels:
            return

        current_level = "L" + level_in_val.group(1)
        if current_level in spec_levels:
            return  # 等级正确

        # 尝试用spec_levels中的等级替换
        for spec_level in sorted(spec_levels):
            candidate = val.replace(current_level, spec_level)
            if candidate in valid:
                fix_log.append({
                    "field": field,
                    "old": val,
                    "new": candidate,
                    "ratio": 90,
                    "source": "等级代码修正"
                })
                cond["value"] = candidate
                return

    def fix_config(config):
        for cond in config.get("common_denominator", {}).get("conditions", []):
            fix_cond(cond)
        for metric in config.get("metrics", []):
            for cond in metric.get("numerator_conditions", []):
                fix_cond(cond)
            for group in metric.get("numerator_or_conditions", []):
                for cond in group:
                    fix_cond(cond)
            for cond in metric.get("custom_denominator_conditions", []):
                fix_cond(cond)

    fix_config(parsed_config)
    return parsed_config, fix_log


def migrate_or_conditions_to_flat(parsed_config: dict) -> dict:
    """将 numerator_or_conditions 迁移到 numerator_conditions 的平铺格式

    AI 解析时把 OR 条件放在 numerator_or_conditions（每组一个条件），
    新版 UI 只有一个统一条件列表，因此需要在加载时做迁移：
    - 把各 OR 组的条件全部并入 numerator_conditions
    - 清空 numerator_or_conditions
    - 保持 numerator_logic = "or"
    """
    for metric in parsed_config.get("metrics", []):
        or_groups = metric.get("numerator_or_conditions", [])
        if or_groups and metric.get("numerator_logic") == "or":
            flat = list(metric.get("numerator_conditions", []))
            for group in or_groups:
                flat.extend(group)
            metric["numerator_conditions"] = flat
            metric["numerator_or_conditions"] = []
    return parsed_config


def extract_group_dimensions_from_spec(spec_bytes, df_columns):
    """从口径Excel的表头结构中提取分组维度

    规范：第1行合并单元格 = 分组字段名��第2行 = 该字段下的具体值
    "全部数据"、"汇总"等总计列不算分组维度，跳过。

    Returns: [{"label": "手指状态", "field": "手指状态", "values": ["手持物", ...]}, ...]
    """
    import difflib
    from openpyxl import load_workbook

    skip_keywords = ["全部", "汇总", "口径", "说明", "统计", "分项", "总计", "合计"]

    try:
        wb = load_workbook(io.BytesIO(spec_bytes), data_only=True)
        ws = wb.active
    except Exception:
        return []

    if ws.max_row < 2:
        return []

    # 收集第1行的合并单元格信息
    merged_in_row1 = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == 1 and merged_range.max_row == 1 and merged_range.max_col > merged_range.min_col:
            val = ws.cell(1, merged_range.min_col).value
            if val is not None:
                val = str(val).strip()
                if val and not any(kw in val for kw in skip_keywords):
                    merged_in_row1.append({
                        "label": val,
                        "min_col": merged_range.min_col,
                        "max_col": merged_range.max_col
                    })

    # 也检查第1行非合并但有值的单元格（单列分组）
    merged_cols = set()
    for m in merged_in_row1:
        for c in range(m["min_col"], m["max_col"] + 1):
            merged_cols.add(c)

    # 对每个合并区域，读取第2行对应列的值作为子项
    results = []
    for m in merged_in_row1:
        values = []
        for c in range(m["min_col"], m["max_col"] + 1):
            v = ws.cell(2, c).value
            if v is not None:
                v_str = str(v).strip()
                if v_str and not any(kw in v_str for kw in skip_keywords):
                    values.append(v_str)
        if not values:
            continue

        # 匹配df_columns
        label = m["label"]
        field = None
        if label in df_columns:
            field = label
        else:
            match = difflib.get_close_matches(label, df_columns, n=1, cutoff=0.6)
            if match:
                field = match[0]

        results.append({"label": label, "field": field, "values": values})

    return results


def extract_spec_descriptions(spec_bytes):
    """从口径Excel提取每行指标的描述列（统计维度、分项内容、口径说明等）

    Returns: {指标名: {"统计维度": "...", "分项内容": "...", "口径说明": "..."}, ...}
    """
    from openpyxl import load_workbook

    desc_col_names = ["统计维度", "分项内容", "口径说明"]

    try:
        wb = load_workbook(io.BytesIO(spec_bytes), data_only=True)
        ws = wb.active
    except Exception:
        return {}

    if ws.max_row < 2:
        return {}

    # 展开合并单元格
    for merged_range in list(ws.merged_cells.ranges):
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        top_left_value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).value = top_left_value

    # 在前两行中查找描述列的位置
    header_rows = []
    for r in range(1, min(3, ws.max_row + 1)):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            row_vals.append(str(v).strip() if v is not None else "")
        header_rows.append(row_vals)

    # 找到描述列在哪一行、哪一列
    desc_col_map = {}  # {列名: 列索引(0-based)}
    header_row_idx = -1
    for row_idx, row_vals in enumerate(header_rows):
        for col_idx, val in enumerate(row_vals):
            for dc in desc_col_names:
                if dc in val and dc not in desc_col_map:
                    desc_col_map[dc] = col_idx
                    header_row_idx = max(header_row_idx, row_idx)

    if not desc_col_map:
        return {}

    # 找"分项内容"列作为指标名来源；如果没有，用第一个描述列右边的列
    indicator_col = desc_col_map.get("分项内容")

    # 数据从 header_row_idx+1 行之后开始（0-based转1-based: +2）
    data_start_row = header_row_idx + 2  # 1-based row number

    result = {}
    for r in range(data_start_row, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            row_vals.append(str(v).strip() if v is not None else "")

        # 获取指标名
        indicator_name = ""
        if indicator_col is not None and indicator_col < len(row_vals):
            indicator_name = row_vals[indicator_col]

        if not indicator_name:
            continue

        desc = {}
        for col_name, col_idx in desc_col_map.items():
            if col_idx < len(row_vals):
                desc[col_name] = row_vals[col_idx]
        result[indicator_name] = desc

    return result


def render_field_mappings(parsed_result, df_columns):
    """展示AI解析的字段映射确认表，低置信度的允许手动修正"""
    mappings = parsed_result.get("field_mappings", [])
    if not mappings:
        return

    st.markdown("**字段映射确认**")
    st.caption("置信度 < 80 的映射请检查并手动修正，修正后重新确认口径")

    has_low_confidence = False
    corrected_mappings = []

    for i, m in enumerate(mappings):
        conf = m.get("confidence", 100)
        if conf >= 80:
            icon = "✅"
        elif conf >= 60:
            icon = "⚠️"
            has_low_confidence = True
        else:
            icon = "🔴"
            has_low_confidence = True

        col1, col2, col3, col4 = st.columns([2, 2, 1, 1])
        with col1:
            st.text(f"{icon} {m.get('description', '')}")
        with col2:
            if conf < 80:
                corrected_field = st.selectbox(
                    "字段",
                    options=df_columns,
                    index=df_columns.index(m["field"]) if m.get("field") in df_columns else 0,
                    key=f"mapping_field_{i}",
                    label_visibility="collapsed"
                )
            else:
                st.text(m.get("field", ""))
                corrected_field = m.get("field", "")
        with col3:
            st.text(m.get("operator", m.get("op", "")))
        with col4:
            st.text(f"{conf}%")

        corrected_mappings.append({**m, "field": corrected_field})

    if has_low_confidence:
        st.warning("存在低置信度映射，请确认上方字段选择正确后再执行统计")

    return corrected_mappings


def read_excel_cached(file_bytes: bytes, header_index: Optional[int]):
    return pd.read_excel(io.BytesIO(file_bytes), header=header_index)


@st.dialog("口径配置详情", width="large")
def show_config_dialog(config: dict):
    import json
    config_str = json.dumps(config, ensure_ascii=False, indent=2)
    st.code(config_str, language="json")
    st.button("📋 复制到剪贴板", on_click=lambda: st.write(
        f'<script>navigator.clipboard.writeText({json.dumps(config_str)})</script>',
        unsafe_allow_html=True
    ), key="copy_config_btn")
    st.caption("提示：点击代码块右上角图标也可复制")


def read_spec_excel(file_bytes: bytes) -> str:
    """读取口径Excel，处理合并单元格和多层表头，返回结构化文本供模型解析。

    多层表头（如第1行"手指状态"合并跨列，第2行为"手持物|手指交叉|..."）
    会被合并为 "手指状态/手持物"、"手指状态/手指交叉" 等层级列名。
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # 记录哪些单元格属于横向合并（跨列），用于检测多层表头
    col_merged_rows = set()  # 含有横向合并单元格的行号
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_col > merged_range.min_col:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                col_merged_rows.add(r)

    # 展开合并单元格，用左上角值填充所有格子
    for merged_range in list(ws.merged_cells.ranges):
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        top_left_value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).value = top_left_value

    # 读取所有行
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(v).strip() if v is not None else "" for v in row]
        all_rows.append(cells)

    if not all_rows:
        return ""

    # 检测多层表头：只检查前3行，需同时满足：有横向合并 + 内容为短文本（<=15字）
    max_header_scan = min(3, len(all_rows))
    header_row_count = 0
    for i in range(max_header_scan):
        row_num = i + 1  # openpyxl行号从1开始
        if row_num not in col_merged_rows:
            break
        # 检查该行是否全是短文本（表头特征，长文本是数据描述不是表头）
        non_empty_cells = [c for c in all_rows[i] if c]
        if non_empty_cells and all(len(c) <= 15 for c in non_empty_cells):
            header_row_count = row_num
        else:
            break

    # 如果检测到多层表头（>=2行），合并为层级列名
    if header_row_count >= 2:
        header_layers = all_rows[:header_row_count]
        data_rows = all_rows[header_row_count:]

        # 合并多层表头：逐列拼接非空的层级名，用 / 分隔
        num_cols = max(len(r) for r in header_layers) if header_layers else 0
        combined_headers = []
        for col_idx in range(num_cols):
            parts = []
            prev_part = None
            for layer in header_layers:
                val = layer[col_idx] if col_idx < len(layer) else ""
                # 跳过空值和与上层重复的值（合并单元格纵向填充导致）
                if val and val != prev_part:
                    parts.append(val)
                prev_part = val
            combined_headers.append("/".join(parts) if parts else "")

        # 输出：第一行为合并后的表头，后续为数据行
        lines = [" | ".join(combined_headers)]
        for row in data_rows:
            if any(row):
                # 补齐列数
                padded = row + [""] * (num_cols - len(row)) if len(row) < num_cols else row
                lines.append(" | ".join(padded[:num_cols]))

        return "\n".join(lines)

    # 无多层表头，按原逻辑输出
    lines = []
    for row in all_rows:
        if any(row):
            lines.append(" | ".join(row))

    return "\n".join(lines)


def get_provider_presets():
    return {
        "自定义": {"base_url": "", "model_name": ""},
        "智谱 GLM": {"base_url": "https://open.bigmodel.cn/api/paas/v4", "model_name": "glm-4"},
        "阿里云百炼（OpenAI兼容）": {"base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1", "model_name": "qwen-plus"},
        "OpenAI": {"base_url": "https://api.openai.com/v1", "model_name": "gpt-4o-mini"}
    }


def render_condition_editor(df, condition_list: list, prefix: str, columns: list, key_prefix: str, conditions_key: str = None):
    """渲染条件编辑器

    Args:
        conditions_key: 可选，用于删除回调。如果不传，删除后不会自动同步到 session_state
    """
    if not condition_list:
        st.info(f"暂无{prefix}条件")
        return

    field_dist = DataService(df).build_field_distribution()

    i = 0
    while i < len(condition_list):
        condition = condition_list[i]
        col1, col2, col3, col4 = st.columns([3, 2, 3, 1])

        with col1:
            current_field = condition.get("field", "")
            field_options = columns if columns else ["（暂无可用字段）"]
            field_index = field_options.index(current_field) if current_field in field_options else 0

            new_field = st.selectbox(
                f"字段",
                options=field_options,
                index=field_index,
                key=f"{key_prefix}_field_{i}",
                disabled=not columns
            )
            if not columns:
                new_field = ""

        with col2:
            op_options = ["等于", "不等于", "包含", "不包含", "为空", "不为空", "属于", "不属于", "大于", "小于"]
            op_values = ["==", "!=", "contains", "not_contains", "is_empty", "is_not_empty", "in", "not_in", "greater_than", "less_than"]
            current_op = condition.get("op", "==")
            op_index = op_values.index(current_op) if current_op in op_values else 0

            selected_op_label = st.selectbox(
                f"运算符",
                options=op_options,
                index=op_index,
                key=f"{key_prefix}_op_{i}"
            )
            new_op = op_values[op_options.index(selected_op_label)]

        with col3:
            if new_op in ["is_empty", "is_not_empty"]:
                new_value = ""
                st.text_input(f"值", value="(无需填写)", disabled=True, key=f"{key_prefix}_value_{i}")
            elif new_op in ["greater_than", "less_than"]:
                current_value = condition.get("value", "")
                new_value = st.text_input(
                    f"值",
                    value=current_value,
                    key=f"{key_prefix}_value_{i}",
                    placeholder="请输入数值"
                )
            elif new_op in ["in", "not_in"]:
                current_value = condition.get("value", "")
                new_value = st.text_input(
                    f"值",
                    value=current_value,
                    key=f"{key_prefix}_value_{i}",
                    placeholder="多个值用英文逗号分隔，如: A,B,C"
                )
            else:
                # 检查该字段是否含多选值，优先展示拆分后的独立值
                field_info = field_dist.get(new_field, {})
                current_value = condition.get("value", "")

                if field_info.get("type") == "categorical_with_multi":
                    value_options = field_info.get("options", [])
                else:
                    value_options = get_column_unique_values(df, new_field)

                if value_options:
                    new_value = st.selectbox(
                        f"值",
                        options=value_options,
                        index=value_options.index(current_value) if current_value in value_options else 0,
                        key=f"{key_prefix}_value_{i}"
                    )
                else:
                    new_value = st.text_input(
                        f"值",
                        value=current_value,
                        key=f"{key_prefix}_value_{i}",
                        placeholder="该字段暂无可选值，请手动输入"
                    )

        with col4:
            if conditions_key:
                # 使用 on_click 回调，避免 fragment 内按钮状态问题
                st.button("🗑️", key=f"{key_prefix}_del_{i}", help=f"删除此条件",
                          on_click=_cb_delete_cond, args=(conditions_key, i))
            else:
                # 兼容旧调用方式：删除后跳出循环，等待下次渲染
                if st.button("🗑️", key=f"{key_prefix}_del_{i}", help=f"删除此条件"):
                    condition_list.pop(i)
                    return  # 删除后直接返回，触发重新渲染

        # 再次检查索引有效性（防止删除后索引越界）
        if i < len(condition_list):
            condition_list[i] = {"field": new_field, "op": new_op, "value": new_value}
        i += 1

    return condition_list


def add_condition(condition_list: list, default_field: str = None):
    """添加新条件"""
    new_condition = {"field": default_field or "", "op": "==", "value": ""}
    condition_list.append(new_condition)


def validate_metrics_config(config: dict, columns: list) -> tuple:
    """验证口径配置是否有效"""
    errors = []
    valid_fields = set(columns)

    for cond in config.get("common_denominator", {}).get("conditions", []):
        if cond.get("field") not in valid_fields:
            errors.append(f"公共分母中使用了无效字段: {cond.get('field')}")

    for i, metric in enumerate(config.get("metrics", [])):
        metric_name = metric.get("name", f"指标{i+1}")

        for cond in metric.get("numerator_conditions", []):
            if cond.get("field") not in valid_fields:
                errors.append(f"[{metric_name}] 分子条件使用了无效字段: {cond.get('field')}")

        if metric.get("denominator_type") == "custom":
            for cond in metric.get("custom_denominator_conditions", []):
                if cond.get("field") not in valid_fields:
                    errors.append(f"[{metric_name}] 自定义分母条件使用了无效字段: {cond.get('field')}")

    return len(errors) == 0, errors


def color_percentage(val):
    """根据百分比值着色"""
    if pd.isna(val) or val is None:
        return 'color: #64748b'
    elif val >= 90:
        return 'color: #2f7a56; font-weight: 600'
    elif val >= 60:
        return 'color: #8b6a2b; font-weight: 600'
    else:
        return 'color: #a63f3f; font-weight: 600'


def color_gap(val):
    """根据gap值着色"""
    if pd.isna(val) or val is None:
        return 'color: #64748b'
    elif val > 0:
        return 'color: #2f7a56; font-weight: 600'
    elif val < 0:
        return 'color: #a63f3f; font-weight: 600'
    else:
        return 'color: #64748b'


def derive_case_filters_from_metrics(metrics_config: dict):
    """从确认的口径配置中提取 badcase/goodcase 预设条件"""
    badcase_presets = []
    for metric in metrics_config.get("metrics", []):
        name = metric.get("name", "")
        numer_conds = metric.get("numerator_conditions", [])
        if numer_conds:
            badcase_presets.append({
                "name": name,
                "conditions": numer_conds
            })
    return badcase_presets, []


@st.fragment
def render_case_filter_tab(df, case_type: str, conditions_key: str, result_key: str):
    """
    渲染Case筛选Tab

    Args:
        df: DataFrame
        case_type: "badcase" 或 "goodcase"
        conditions_key: session_state中存储条件的key
        result_key: session_state中存储结果的key
    """
    import copy  # 移到函数开头，避免 UnboundLocalError
    import json

    conditions = st.session_state.get(conditions_key, [])

    # 从口径配置自动生成的预设条件
    presets = st.session_state.get(f"{case_type}_presets", [])
    if presets:
        st.markdown("**快速填入（从口径配置自动生成）**")
        preset_names = ["（手动配置）"] + [p["name"] for p in presets]
        selected = st.selectbox("选择指标条件", preset_names, key=f"{case_type}_preset_select")
        # 只在切换预设时自动填入，避免覆盖用户后续的手动编辑
        last_applied_key = f"{case_type}_last_applied_preset"
        last_applied = st.session_state.get(last_applied_key, "")
        if selected != "（手动配置）":
            if selected != last_applied:
                preset = next((p for p in presets if p["name"] == selected), None)
                if preset:
                    st.session_state[conditions_key] = copy.deepcopy(preset["conditions"])
                    st.session_state[last_applied_key] = selected
                    st.rerun()
        else:
            if last_applied != "":
                st.session_state[last_applied_key] = ""

    # 条件配置区
    st.markdown("**筛选条件**（多条件AND关系）")

    if conditions:
        render_condition_editor(
            df,
            conditions,
            case_type,
            st.session_state["columns"],
            f"{case_type}_cond",
            conditions_key=conditions_key  # 传入用于删除回调
        )

    col1, col2, col3 = st.columns([1, 1, 3])

    with col1:
        st.button("➕ 添加条件", key=f"add_{case_type}_cond",
                  use_container_width=True,
                  on_click=_cb_add_case_cond, args=(conditions_key,))

    with col2:
        st.button("🗑️ 清空条件", key=f"clear_{case_type}_cond", use_container_width=True,
                  on_click=_cb_clear_case_conds, args=(conditions_key, result_key))

    # 更新条件
    st.session_state[conditions_key] = conditions

    # 实时过滤：条件变化时自动执行，无需点击按钮
    prev_conditions_key = f"{conditions_key}_prev"
    try:
        conditions_sig = json.dumps(conditions, ensure_ascii=False, sort_keys=True)
        prev_sig = json.dumps(st.session_state.get(prev_conditions_key, []), ensure_ascii=False, sort_keys=True)
    except Exception:
        conditions_sig, prev_sig = str(conditions), ""

    if conditions_sig != prev_sig:
        if conditions:
            result_df = DataService(df).filter_cases(conditions)
            st.session_state[result_key] = result_df
        else:
            st.session_state[result_key] = None
        st.session_state[prev_conditions_key] = copy.deepcopy(conditions)

    st.divider()

    # 显示筛选结果
    result_df = st.session_state.get(result_key)

    if result_df is not None and not result_df.empty:
        total_count = len(df)
        filtered_count = len(result_df)

        st.success(f"筛选完成：**{filtered_count}** 条 / 共 **{total_count}** 条")

        # 数据预览
        st.markdown("**筛选结果预览**（前50行）")

        # 选择显示的列
        display_columns = st.multiselect(
            "选择显示列",
            options=st.session_state["columns"],
            default=st.session_state["columns"][:5] if len(st.session_state["columns"]) > 5 else st.session_state["columns"],
            key=f"display_cols_{case_type}"
        )

        if display_columns:
            preview_df = result_df[display_columns].head(50)
            st.dataframe(preview_df, use_container_width=True, hide_index=True)
        else:
            st.dataframe(result_df.head(50), use_container_width=True, hide_index=True)

        # 单条数据详情展开
        with st.expander("📋 查看单条数据详情"):
            detail_index = st.number_input(
                "选择行号（从0开始）",
                min_value=0,
                max_value=len(result_df) - 1,
                value=0,
                key=f"detail_idx_{case_type}"
            )

            if detail_index < len(result_df):
                detail_row = result_df.iloc[detail_index]
                for col in result_df.columns:
                    st.markdown(f"**{col}**: {detail_row[col]}")

        # 导出按钮
        st.divider()

        col1, col2 = st.columns([1, 3])

        with col1:
            with st.spinner("处理中..."):
                export_path = ExportService().export_filtered_cases_to_excel(
                    result_df,
                    "评测结果",
                    case_type
                )

            with open(export_path, "rb") as f:
                st.download_button(
                    label=f"导出为Excel（{filtered_count}条）",
                    data=f,
                    file_name=f"{case_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"export_{case_type}_btn"
                )

    elif result_df is not None and result_df.empty:
        st.warning("没有符合条件的数据")

    else:
        st.info("👆 请添加筛选条件，结果将自动更新")


def _normalize_metric_name(name: str) -> str:
    if not name:
        return ""
    normalized = str(name).strip().lower()
    normalized = re.sub(r"\s+", "", normalized)
    normalized = re.sub(r"[：:（）()【】\[\]，,。\.·\-_/]", "", normalized)
    replacements = {
        "服装整体": "服装",
        "整体": "",
        "总体": "",
        "异常率": "不还原率"
    }
    for old, new in replacements.items():
        normalized = normalized.replace(old, new)
    return normalized


def _find_best_match_name(spec_name, results_map):
    if spec_name in results_map:
        return spec_name

    spec_norm = _normalize_metric_name(spec_name)
    if not spec_norm:
        return None

    normalized_name_map = {}
    for rname in results_map:
        r_norm = _normalize_metric_name(rname)
        if r_norm and r_norm not in normalized_name_map:
            normalized_name_map[r_norm] = rname

    if spec_norm in normalized_name_map:
        return normalized_name_map[spec_norm]

    for rname in results_map:
        r_norm = _normalize_metric_name(rname)
        if spec_norm in r_norm or r_norm in spec_norm:
            return rname

    spec_has_negation = "不" in spec_norm
    best_name = None
    best_ratio = 0.0
    for rname in results_map:
        r_norm = _normalize_metric_name(rname)
        if not r_norm:
            continue
        if ("不" in r_norm) != spec_has_negation:
            continue
        ratio = difflib.SequenceMatcher(None, spec_norm, r_norm).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_name = rname
    if best_name and best_ratio >= 0.72:
        return best_name
    return None


def _fuzzy_match(spec_name, results_map):
    matched_name = _find_best_match_name(spec_name, results_map)
    if matched_name:
        return results_map.get(matched_name, {})
    return {}


def build_stats_table_from_spec(stats_result, spec_descriptions, is_grouping=False, grouping_result=None):
    """以口径Excel结构为基准构建统计表，未匹配的行保留空值"""
    rows = []
    results_map = {r["name"]: r for r in stats_result.get("results", [])}
    denom_count = stats_result.get("denominator_count", 0)

    for metric_name, desc_info in spec_descriptions.items():
        row = dict(desc_info)
        row["指标名称"] = metric_name
        matched = _fuzzy_match(metric_name, results_map)

        if is_grouping and grouping_result is not None:
            groups = grouping_result.get("groups", [])
            for g in groups:
                g_data = matched.get(g, {})
                pct = g_data.get("percentage")
                numer = g_data.get("numerator", 0)
                denom = g_data.get("denominator", 0)
                row[g] = f"{numer}/{denom} ({pct}%)" if pct is not None else f"N/A (0/0)"
        else:
            row["分子"] = matched.get("numerator", "")
            row["分母"] = matched.get("denominator", denom_count) if matched else ""
            pct = matched.get("percentage")
            row["百分比"] = f"{pct}%" if pct is not None else ""
            if matched:
                row["原始百分比"] = pct

        # "评测总数"行：从 denominator_count 填入
        if not matched and ("评测总数" in metric_name or "总数" in metric_name):
            if is_grouping and grouping_result is not None:
                denom_counts = grouping_result.get("denominator_counts", {})
                total_count = grouping_result.get("denominator_count", 0)
                groups = grouping_result.get("groups", [])
                for g in groups:
                    g_count = denom_counts.get(g, 0)
                    if g == "全部":
                        row[g] = str(g_count)
                    elif total_count > 0:
                        pct = round(g_count / total_count * 100, 2)
                        row[g] = f"{g_count}/{total_count} ({pct}%)"
                    else:
                        row[g] = str(g_count)
            else:
                row["分子"] = denom_count
                row["分母"] = denom_count
                row["百分比"] = ""

        rows.append(row)

    return pd.DataFrame(rows)


# ==================== Fragment回调函数（避免rerun滚动跳动） ====================

def _cb_add_denom_cond():
    editing = st.session_state.get("editing_metrics", {})
    conds = editing.get("common_denominator", {}).get("conditions", [])
    cols = st.session_state.get("columns", [])
    add_condition(conds, cols[0] if cols else None)
    editing.setdefault("common_denominator", {})["conditions"] = conds
    st.session_state["editing_metrics"] = editing


def _cb_use_ai_denom():
    """一键将AI分母条件复制到手动配置conditions"""
    import copy
    editing = st.session_state.get("editing_metrics", {})
    ai_conds = editing.get("common_denominator", {}).get("ai_analyzed_conditions", [])
    editing.setdefault("common_denominator", {})["conditions"] = copy.deepcopy(ai_conds)
    st.session_state["editing_metrics"] = editing


def _cb_clear_denom():
    """清空手动配置conditions（全部数据）"""
    editing = st.session_state.get("editing_metrics", {})
    editing.setdefault("common_denominator", {})["conditions"] = []
    st.session_state["editing_metrics"] = editing


def _cb_add_denom_group():
    """添加一个新的公共分母条件组"""
    editing = st.session_state.get("editing_metrics", {})
    groups = editing.get("common_denominator", {}).get("condition_groups", [])
    groups.append({"name": f"条件组{len(groups)+1}", "conditions": []})
    editing.setdefault("common_denominator", {})["condition_groups"] = groups
    st.session_state["editing_metrics"] = editing


def _cb_del_denom_group(gi):
    """删除指定索引的公共分母条件组"""
    editing = st.session_state.get("editing_metrics", {})
    groups = editing.get("common_denominator", {}).get("condition_groups", [])
    if gi < len(groups):
        groups.pop(gi)
    editing.setdefault("common_denominator", {})["condition_groups"] = groups
    st.session_state["editing_metrics"] = editing


def _cb_add_denom_group_cond(gi):
    """向指定组添加一个空条件"""
    editing = st.session_state.get("editing_metrics", {})
    groups = editing.get("common_denominator", {}).get("condition_groups", [])
    if gi < len(groups):
        cols = st.session_state.get("columns", [])
        add_condition(groups[gi]["conditions"], cols[0] if cols else None)
    editing.setdefault("common_denominator", {})["condition_groups"] = groups
    st.session_state["editing_metrics"] = editing


def _cb_add_numer_cond(metric_idx):
    editing = st.session_state.get("editing_metrics", {})
    metrics = editing.get("metrics", [])
    if metric_idx < len(metrics):
        conds = metrics[metric_idx].get("numerator_conditions", [])
        cols = st.session_state.get("columns", [])
        add_condition(conds, cols[0] if cols else None)
        metrics[metric_idx]["numerator_conditions"] = conds
    st.session_state["editing_metrics"] = editing


def _cb_add_custom_denom_cond(metric_idx):
    editing = st.session_state.get("editing_metrics", {})
    metrics = editing.get("metrics", [])
    if metric_idx < len(metrics):
        conds = metrics[metric_idx].get("custom_denominator_conditions", [])
        cols = st.session_state.get("columns", [])
        add_condition(conds, cols[0] if cols else None)
        metrics[metric_idx]["custom_denominator_conditions"] = conds
    st.session_state["editing_metrics"] = editing


def _cb_add_case_cond(conditions_key):
    """Case筛选：添加一个空条件"""
    conditions = st.session_state.get(conditions_key, [])
    cols = st.session_state.get("columns", [])
    add_condition(conditions, cols[0] if cols else None)
    st.session_state[conditions_key] = conditions


def _cb_delete_cond(conditions_key, index):
    """删除指定索引的条件"""
    conditions = st.session_state.get(conditions_key, [])
    if 0 <= index < len(conditions):
        conditions.pop(index)
        st.session_state[conditions_key] = conditions


def _cb_clear_case_conds(conditions_key, result_key):
    """清空条件"""
    st.session_state[conditions_key] = []
    st.session_state[result_key] = None


def _cb_add_metric():
    editing = st.session_state.get("editing_metrics", {})
    metrics = editing.setdefault("metrics", [])
    metrics.append({
        "name": f"新指标{len(metrics) + 1}",
        "numerator_conditions": [],
        "denominator_type": "common",
        "custom_denominator_conditions": []
    })
    st.session_state["editing_metrics"] = editing


@st.fragment
def render_metrics_editor_fragment():
    """口径编辑区域 fragment，内部操作通过 on_click 回调避免 rerun 滚动跳动"""
    editing = st.session_state.get("editing_metrics", {})
    if not editing:
        return

    # 公共分母编辑
    st.markdown("### 📊 公共分母")
    if "common_denominator" not in editing:
        editing["common_denominator"] = {"description": "", "conditions": [], "type": "ai", "ai_analyzed_conditions": []}

    editing["common_denominator"]["description"] = st.text_input("分母描述", value=editing["common_denominator"].get("description", ""), key="denom_desc")

    # 获取 AI 分析的分母条件
    ai_denom_conds = editing["common_denominator"].get("ai_analyzed_conditions", [])
    has_ai_parsed = "ai_analyzed_conditions" in editing.get("common_denominator", {})
    manual_conditions = editing["common_denominator"].get("conditions", [])

    # 首页两个选项：AI分析的分母 / 手动配置
    denom_type_options = ["AI分析的分母", "手动配置"]
    denom_type_values = ["ai", "custom"]

    current_type = editing["common_denominator"].get("type", "ai")
    # 如果没有AI解析过且当前是ai类型，降级为custom
    if current_type == "ai" and not has_ai_parsed:
        current_type = "custom"
    if current_type not in denom_type_values:
        current_type = "custom"
    type_idx = denom_type_values.index(current_type)

    selected_type = st.radio("分母类型", denom_type_options, index=type_idx,
                             key="denom_type_radio", horizontal=True)
    editing["common_denominator"]["type"] = denom_type_values[denom_type_options.index(selected_type)]
    denom_type = editing["common_denominator"]["type"]

    # 根据类型显示不同内容
    if denom_type == "ai":
        # AI分析的分母：显示可编辑的条件
        st.markdown("**AI分析的分母条件**")
        if ai_denom_conds:
            # AI条件可编辑
            render_condition_editor(st.session_state["df"], ai_denom_conds, "AI分母", st.session_state["columns"], "ai_denom_cond")
            # 同步更新
            editing["common_denominator"]["ai_analyzed_conditions"] = ai_denom_conds
        else:
            st.info("AI未解析出分母条件，等同于「全部数据」。")

    else:  # custom - 手动配置
        # 快捷按钮区域
        btn_col1, btn_col2, _ = st.columns([1, 1, 2])
        with btn_col1:
            if has_ai_parsed:
                st.button("🤖 使用AI分母", key="use_ai_denom_btn",
                          help="一键使用AI分析的分母条件",
                          use_container_width=True,
                          on_click=_cb_use_ai_denom)
        with btn_col2:
            if manual_conditions:
                st.button("📊 全部数据", key="clear_denom_btn",
                          help="清空条件，使用全部数据",
                          use_container_width=True,
                          on_click=_cb_clear_denom)

        # 条件组管理区
        condition_groups = editing["common_denominator"].get("condition_groups", [])
        st.markdown("**条件组**（指标自定义分母可以选择不同组）")

        for gi, grp in enumerate(condition_groups):
            grp_name = grp.get("name", f"条件组{gi+1}") if isinstance(grp, dict) else grp.name
            grp_conds = grp.get("conditions", []) if isinstance(grp, dict) else grp.conditions

            with st.expander(f"📁 {grp_name}", expanded=True):
                g_col1, g_col2 = st.columns([4, 1])
                with g_col1:
                    new_name = st.text_input("组名", value=grp_name, key=f"denom_grp_name_{gi}", label_visibility="collapsed")
                    if isinstance(grp, dict):
                        grp["name"] = new_name
                with g_col2:
                    st.button("🗑️", key=f"del_denom_grp_{gi}",
                              on_click=_cb_del_denom_group, args=(gi,),
                              help="删除此组")
                if grp_conds:
                    render_condition_editor(st.session_state["df"], grp_conds, f"分母组{gi}", st.session_state["columns"], f"denom_grp_cond_{gi}")
                    if isinstance(grp, dict):
                        grp["conditions"] = grp_conds
                else:
                    st.caption("此组还没有条件，点击下方「添加条件」。")
                st.button("➕ 添加条件", key=f"add_denom_grp_cond_{gi}",
                          on_click=_cb_add_denom_group_cond, args=(gi,))

        editing["common_denominator"]["condition_groups"] = condition_groups
        st.button("➕ 添加条件组", key="add_denom_grp", on_click=_cb_add_denom_group)

        # ���用条件（不属于任何组，作为全局手动条件保留兼容）
        if manual_conditions:
            st.markdown("**通用条件**（不分组）")
            render_condition_editor(st.session_state["df"], manual_conditions, "分母", st.session_state["columns"], "denom_cond")
        if st.button("➕ 添加通用条件", key="add_denom_cond", on_click=_cb_add_denom_cond):
            pass

    # 实时预览公共分母匹配数量
    from services_2_1 import DataService
    df_preview = st.session_state["df"]
    manual_conditions = editing["common_denominator"].get("conditions", [])
    if denom_type == "ai" and ai_denom_conds:
        df_denom_preview = DataService(df_preview).apply_conditions(ai_denom_conds)
    elif denom_type == "custom" and manual_conditions:
        df_denom_preview = DataService(df_preview).apply_conditions(manual_conditions)
    else:
        df_denom_preview = df_preview
    st.info(f"当前公共分母匹配数据量: **{len(df_denom_preview)}** / {len(df_preview)} 条")

    editing["common_denominator"]["conditions"] = manual_conditions

    st.divider()

    # 指标列表编辑
    st.markdown("### 📈 指标列表")
    if "metrics" not in editing:
        editing["metrics"] = []

    metrics_to_remove = []

    for i, metric in enumerate(editing["metrics"]):
        with st.expander(f"**{metric.get('name', f'指标{i+1}')}**", expanded=True):
            col1, col2 = st.columns([5, 1])

            with col1:
                metric["name"] = st.text_input("指标名称", value=metric.get("name", f"指标{i+1}"), key=f"metric_name_{i}")

            with col2:
                if st.button("🗑️ 删除指标", key=f"del_metric_{i}"):
                    metrics_to_remove.append(i)

            # 分子条件 AND/OR 模式切换
            logic_options = ["全部满足 (AND)", "任一满足 (OR)"]
            logic_values = ["and", "or"]
            current_logic = metric.get("numerator_logic", "and")
            logic_idx = logic_values.index(current_logic) if current_logic in logic_values else 0
            selected_logic = st.radio("分子条件关系", logic_options, index=logic_idx,
                                      key=f"numer_logic_{i}", horizontal=True)
            metric["numerator_logic"] = logic_values[logic_options.index(selected_logic)]

            if metric["numerator_logic"] == "or":
                st.markdown("**分子条件（OR，任一满足即计入）**")
            else:
                st.markdown("**分子条件（AND，所有条件同时满足）**")
            numer_conditions = metric.get("numerator_conditions", [])

            if numer_conditions:
                render_condition_editor(st.session_state["df"], numer_conditions, "分子", st.session_state["columns"], f"numer_cond_{i}")

            st.button("➕ 添加条件", key=f"add_numer_{i}",
                      on_click=_cb_add_numer_cond, args=(i,))

            metric["numerator_conditions"] = numer_conditions


            st.markdown("**分母选择**")
            denom_options = ["使用公共分母", "自定义分母"]
            denom_values = ["common", "custom"]
            current_denom = metric.get("denominator_type", "common")
            denom_index = denom_values.index(current_denom) if current_denom in denom_values else 0

            selected_denom_label = st.radio("分母类型", options=denom_options, index=denom_index, key=f"denom_type_{i}", horizontal=True)
            denom_type = denom_values[denom_options.index(selected_denom_label)]
            metric["denominator_type"] = denom_type

            if denom_type == "common":
                # 显示当前公共分母内容
                common_denom = editing.get("common_denominator", {})
                common_type = common_denom.get("type", "ai")
                if common_type == "ai":
                    ai_conds = common_denom.get("ai_analyzed_conditions", [])
                    if ai_conds:
                        cond_strs = [f"{c.get('field')} {c.get('op')} {c.get('value', '')}" for c in ai_conds]
                        st.caption(f"公共分母（AI分析）: {' AND '.join(cond_strs)}")
                    else:
                        st.caption("公共分母（AI分析）: 全部数据")
                else:
                    manual_conds = common_denom.get("conditions", [])
                    if manual_conds:
                        cond_strs = [f"{c.get('field')} {c.get('op')} {c.get('value', '')}" for c in manual_conds]
                        st.caption(f"公共分母（手动配置）: {' AND '.join(cond_strs)}")
                    else:
                        st.caption("公共分母（手动配置）: 全部数据")
                metric["custom_denominator_conditions"] = []
            else:
                # 自定义分母：四选一（AI分母/全部数据/条件组/手动配置）
                st.markdown("**分母来源**")
                cond_groups = editing.get("common_denominator", {}).get("condition_groups", [])
                has_groups = bool(cond_groups)

                source_options = ["AI分析的分母", "全部数据"]
                source_values = ["ai", "all"]
                if has_groups:
                    source_options.append("选择条件组")
                    source_values.append("group")
                source_options.append("手动配置")
                source_values.append("custom")

                # 获取当前自定义分母的来源类型
                custom_denom_conds = metric.get("custom_denominator_conditions", [])
                custom_source = metric.get("custom_denominator_source", "ai")
                if custom_source not in source_values:
                    custom_source = "ai"
                source_idx = source_values.index(custom_source)

                selected_source = st.radio("来源", source_options, index=source_idx, key=f"denom_source_{i}", horizontal=True)
                metric["custom_denominator_source"] = source_values[source_options.index(selected_source)]

                custom_type = metric["custom_denominator_source"]

                if custom_type == "ai":
                    # 使用AI分析的分母：显示可编辑条件
                    ai_conds = editing.get("common_denominator", {}).get("ai_analyzed_conditions", [])
                    if ai_conds:
                        render_condition_editor(st.session_state["df"], ai_conds, f"指标{i}AI分母", st.session_state["columns"], f"metric_ai_denom_{i}")
                        editing["common_denominator"]["ai_analyzed_conditions"] = ai_conds
                    else:
                        st.caption("AI未解析出分母条件，等同于全部数据。")
                    metric["custom_denominator_conditions"] = []

                elif custom_type == "all":
                    st.caption("全部数据（无过滤条件）")
                    metric["custom_denominator_conditions"] = []

                elif custom_type == "group":
                    # 选择条件组
                    group_names = [g.get("name", f"组{gi+1}") if isinstance(g, dict) else g.name for gi, g in enumerate(cond_groups)]
                    current_grp = metric.get("custom_denominator_group", group_names[0] if group_names else "")
                    if current_grp not in group_names and group_names:
                        current_grp = group_names[0]
                    grp_idx = group_names.index(current_grp) if current_grp in group_names else 0
                    selected_grp = st.selectbox("选择条件组", group_names, index=grp_idx, key=f"denom_grp_sel_{i}")
                    metric["custom_denominator_group"] = selected_grp
                    # 显示选中组的条件（只读预览）
                    for g in cond_groups:
                        g_name = g.get("name", "") if isinstance(g, dict) else g.name
                        if g_name == selected_grp:
                            g_conds = g.get("conditions", []) if isinstance(g, dict) else g.conditions
                            if g_conds:
                                cond_strs = [f"{c.get('field')} {c.get('op')} {c.get('value', '')}" for c in g_conds]
                                st.caption(f"条件: {' AND '.join(cond_strs)}")
                            else:
                                st.caption("此组暂无条件")
                            break
                    metric["custom_denominator_conditions"] = []

                else:  # custom - 手动配置
                    if custom_denom_conds:
                        render_condition_editor(st.session_state["df"], custom_denom_conds, "自定义分母", st.session_state["columns"], f"custom_denom_{i}")
                    else:
                        st.caption("💡 点击「添加条件」手动配置分母条件。")
                    st.button("➕ 添加条件", key=f"add_custom_denom_{i}",
                              on_click=_cb_add_custom_denom_cond, args=(i,))
                    metric["custom_denominator_conditions"] = custom_denom_conds

    for i in sorted(metrics_to_remove, reverse=True):
        editing["metrics"].pop(i)

    st.button("➕ 手动添加新指标", key="add_metric", on_click=_cb_add_metric)

    st.session_state["editing_metrics"] = editing


# ==================== 主程序 ====================
init_session_state()

# ==================== 侧边栏 ====================
with st.sidebar:
    st.header("⚙️ API配置")

    provider_presets = get_provider_presets()

    # 显示当前配置状态
    current_provider = st.selectbox(
        "厂商预设",
        options=list(provider_presets.keys()),
        key="provider_preset"
    )

    # 预设详情（默认收起）
    preset = provider_presets[current_provider]
    with st.expander(f"📋 {current_provider} 配置详情", expanded=False):
        st.markdown(f"- **接口地址**: `{preset['base_url'] or '需手动填写'}`")
        st.markdown(f"- **模型名称**: `{preset['model_name'] or '需手动填写'}`")

    if st.button("✅ 应用此预设", use_container_width=True, type="primary"):
        st.session_state["base_url"] = preset["base_url"]
        st.session_state["model_name"] = preset["model_name"]
        st.success(f"已应用【{current_provider}】预设！")

    st.divider()

    # API配置输入
    api_key = st.text_input("API密钥", type="password",
                            placeholder="请输入API密钥",
                            value=st.session_state.get("api_key", os.environ.get("API_KEY", "")),
                            key="api_key")
    base_url = st.text_input("接口地址",
                             placeholder="例如: https://open.bigmodel.cn/api/paas/v4",
                             value=st.session_state.get("base_url", os.environ.get("BASE_URL", "")),
                             key="base_url")
    model_name = st.text_input("模型名称",
                               placeholder="例如: glm-4 / qwen-plus",
                               value=st.session_state.get("model_name", os.environ.get("MODEL_NAME", "")),
                               key="model_name")

    # 显示当前配置状态
    if api_key and base_url:
        st.success("✅ API已配置")
    else:
        st.info("👆 请填写API配置")

    if st.button("🔌 测试API连接", use_container_width=True):
        if api_key and base_url:
            with st.spinner("测试连接中..."):
                try:
                    llm = LLMService(api_key, base_url, model_name or "glm-4")
                    success, message = llm.test_connection()
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
                except Exception as e:
                    st.error(format_api_error(e))
        else:
            st.warning("请先填写API密钥和接口地址")

    col_save, col_del = st.columns(2)
    with col_save:
        if st.button("💾 保存配置", use_container_width=True):
            save_api_config(api_key, base_url, model_name)
            st.success("已保存")
    with col_del:
        if st.button("🗑️ 删除配置", use_container_width=True):
            delete_api_config()
            st.success("已删除")

    with st.expander("📘 如何添加模型API", expanded=False):
        st.markdown("""
        - 填写 **API密钥 / 接口地址 / 模型名称** 即可接入任意兼容 OpenAI 的厂商
        - 常见接口地址：
          - 智谱：`https://open.bigmodel.cn/api/paas/v4`
          - 阿里云：`https://dashscope.aliyuncs.com/compatible-mode/v1`
        - 常见模型：`glm-4`、`qwen-plus`、`gpt-4o-mini`
        """)

    st.divider()

    # 工具按钮
    st.markdown("### 🛠️ 工具")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🔄 重新运行", use_container_width=True):
            st.rerun()
    with col2:
        if st.button("🗑️ 清除缓存", use_container_width=True):
            st.cache_data.clear()
            st.cache_resource.clear()
            st.success("缓存已清除")

    st.divider()

    if st.session_state["df"] is not None:
        st.success(f"✅ 已加载数据: {len(st.session_state['df'])} 行")
    else:
        st.info("📂 请上传数据文件（Excel 或 CSV）")

# ==================== 主界面 ====================
st.title("📊 标注评测报告生成工具")
st.markdown("上传标注Excel → AI解析统计口径 → 人工确认口径 → 自动统计 → 生成报告 → Case筛选")

st.divider()

# ==================== 文件上传区域 ====================
st.subheader("1. 数据上传")

col1, col2 = st.columns([3, 1])

with col1:
    uploaded_file = st.file_uploader(
        "上传标注数据文件",
        type=["xlsx", "xls", "csv"],
        help="支持 .xlsx、.xls、.csv 格式",
        key="file_uploader"
    )

with col2:
    is_csv = uploaded_file is not None and uploaded_file.name.endswith(".csv")
    header_row = st.selectbox(
        "表头在第几行",
        options=list(range(1, 11)),
        index=0,
        help="选择表头所在的行号（从1开始），CSV文件固定为第1行",
        disabled=is_csv
    )

# ==================== 数据处理 ====================
if uploaded_file is not None:
    try:
        with st.spinner("处理中..."):
            uploaded_bytes = uploaded_file.getvalue()
            if uploaded_file.name.endswith(".csv"):
                df_raw = pd.read_csv(io.BytesIO(uploaded_bytes), encoding="utf-8-sig")
            else:
                df_raw = read_excel_cached(uploaded_bytes, header_row - 1)
        st.success(f"文件读取成功：**{uploaded_file.name}**")
        if len(df_raw) > 10000:
            st.info("检测到大文件（>10000行），已启用缓存读取。")

        with st.spinner("正在清洗数据..."):
            ds = DataService(df_raw)
            df_cleaned, clean_report = ds.clean_dataframe()

        # ==================== 清洗报告展示 ====================
        st.subheader("2. 数据清洗报告")

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("原始行数", clean_report["original_rows"])
        with col2:
            st.metric("删除空行", clean_report["empty_rows_removed"])
        with col3:
            st.metric("重复行数", clean_report["duplicate_rows"])
        with col4:
            st.metric("字段数量", len(df_cleaned.columns))

        if clean_report["columns_renamed"]:
            with st.expander("📝 列名已去除空格", expanded=False):
                for item in clean_report["columns_renamed"]:
                    st.write(f"`{item['old']}` → `{item['new']}`")

        null_counts = clean_report["column_null_counts"]
        if any(v > 0 for v in null_counts.values()):
            with st.expander("🔍 各列空值统计", expanded=False):
                null_df = pd.DataFrame([
                    {"字段名": k, "空值数": v, "空值率": f"{v/len(df_cleaned)*100:.1f}%"}
                    for k, v in null_counts.items() if v > 0
                ])
                st.dataframe(null_df, use_container_width=True, hide_index=True)

        if clean_report["duplicate_rows"] > 0:
            st.warning(f"检测到 **{clean_report['duplicate_rows']}** 行完全重复数据")
            remove_dup = st.checkbox("删除重复行", value=False, key="remove_duplicates")
            if remove_dup:
                df_cleaned, removed = remove_duplicates(df_cleaned)
                st.info(f"已删除 {removed} 行重复数据")

        st.subheader("3. 字段总览")
        column_summary = get_column_summary(df_cleaned)
        summary_df = pd.DataFrame(column_summary)
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

        # ==================== 字段分类预览（人工介入点1） ====================
        st.subheader("4. 评估字段确认")
        st.caption("系统自动识别哪些字段是评估标签列，你可以手动调整。只有勾选的字段会发给AI做口径解析。")

        field_dist_raw = DataService(df_cleaned).build_field_distribution()
        # 初始化勾选状态
        if "eval_fields" not in st.session_state or st.session_state.get("_last_columns") != df_cleaned.columns.tolist():
            st.session_state["eval_fields"] = [
                col for col, info in field_dist_raw.items()
                if info["type"] in ("categorical", "categorical_with_multi", "high_cardinality")
            ]
            st.session_state["_last_columns"] = df_cleaned.columns.tolist()

        type_label = {"categorical": "✅ 评估列", "categorical_with_multi": "✅ 多选列", "high_cardinality": "⚠️ 高基数", "skip": "⏭️ 跳过"}
        type_color = {"categorical": "green", "categorical_with_multi": "green", "high_cardinality": "orange", "skip": "gray"}

        with st.expander("查看/调整字段分类", expanded=False):
            cols_per_row = 3
            col_items = list(field_dist_raw.items())
            for row_start in range(0, len(col_items), cols_per_row):
                row_cols = st.columns(cols_per_row)
                for idx, (col, info) in enumerate(col_items[row_start:row_start + cols_per_row]):
                    with row_cols[idx]:
                        label = type_label.get(info["type"], info["type"])
                        checked = col in st.session_state["eval_fields"]
                        new_checked = st.checkbox(
                            f"{label} `{col}`",
                            value=checked,
                            key=f"field_check_{col}",
                            help=f"类型: {info['type']} | " + (
                                f"值: {', '.join(info['values'][:5])}{'...' if len(info.get('values', [])) > 5 else ''}"
                                if info["type"] == "categorical"
                                else f"选项: {', '.join(info['options'][:5])}{'...' if len(info.get('options', [])) > 5 else ''}"
                                if info["type"] == "categorical_with_multi"
                                else f"唯一值数: {info.get('unique_count', '-')}"
                            )
                        )
                        if new_checked and col not in st.session_state["eval_fields"]:
                            st.session_state["eval_fields"].append(col)
                        elif not new_checked and col in st.session_state["eval_fields"]:
                            st.session_state["eval_fields"].remove(col)

        eval_count = len(st.session_state["eval_fields"])
        st.info(f"已选 **{eval_count}** 个评估字段，将作为口径解析的背景信息")

        with st.expander("4. 数据预览", expanded=False):
            preview_rows = st.slider("预览行数", min_value=5, max_value=50, value=10, step=5, key="preview_slider")
            st.dataframe(df_cleaned.head(preview_rows), use_container_width=True, hide_index=True)
            st.caption(f"显示前 {preview_rows} 行，共 {len(df_cleaned)} 行数据")

        st.session_state["df"] = df_cleaned
        st.session_state["columns"] = df_cleaned.columns.tolist()
        st.session_state["clean_report"] = clean_report

        st.divider()

        # ==================== 口径定义区域 ====================
        st.subheader("5. 定义统计口径")

        # 模板类型选择
        template_type = st.radio("模板类型", ["线上评测", "GSB对比", "等级对比"], horizontal=True, key="template_type")

        # GSB模式：模型名配置
        if template_type == "GSB对比":
            st.info("GSB对比模式：口径中直接使用「左边好/右边好」等原始值匹配，模型名仅在报告中替换显示。")
            col_model_a, col_model_b = st.columns(2)
            with col_model_a:
                st.text_input("模型A名称（左边）", value="", placeholder="留空则保持「左边」", key="gsb_model_a")
            with col_model_b:
                st.text_input("模型B名称（右边）", value="", placeholder="留空则保持「右边」", key="gsb_model_b")
        elif template_type == "等级对比":
            st.info("等级对比模式：按 L0/L1/L2/L3 等级标签统计，分母统一，使用与线上评测相同的解析规则。")

        api_configured = bool(st.session_state.get("api_key") and st.session_state.get("base_url"))
        if not api_configured:
            st.warning("⚠️ 请先在左侧配置API密钥和接口地址")

        if "editing_metrics" not in st.session_state or st.session_state["editing_metrics"] is None:
            st.session_state["editing_metrics"] = EMPTY_METRICS_CONFIG.copy()

        # 上传口径Excel
        st.markdown("上传包含口径说明的Excel文件，AI将读取并解析")
        spec_excel = st.file_uploader("上传口径Excel", type=["xlsx", "xls"], key="spec_excel_uploader")

        if spec_excel is not None:
            try:
                spec_bytes = spec_excel.getvalue()
                with st.spinner("处理中..."):
                    spec_content = read_spec_excel(spec_bytes)
                st.text_area("口径内容预览", value=spec_content, height=200, disabled=True, key="spec_excel_preview")

                # 提取分组维度
                group_dims = extract_group_dimensions_from_spec(spec_bytes, st.session_state["columns"])
                if group_dims:
                    st.session_state["auto_group_dimensions"] = group_dims

                # 提取口径描述列
                spec_descs = extract_spec_descriptions(spec_bytes)
                if spec_descs:
                    st.session_state["spec_descriptions"] = spec_descs

                st.button(
                    "解析Excel口径",
                    key="parse_excel_btn",
                    disabled=(not api_configured) or st.session_state.get("trigger_parse_excel", False),
                    on_click=request_parse_excel
                )
                if st.session_state.get("trigger_parse_excel"):
                    st.session_state["trigger_parse_excel"] = False
                    with st.spinner("AI正在解析口径..."):
                        try:
                            clear_metrics_editor_keys()
                            st.session_state["parsed_metrics"] = None
                            st.session_state["editing_metrics"] = None
                            llm = LLMService(st.session_state["api_key"], st.session_state["base_url"], st.session_state.get("model_name", "glm-4"))
                            field_dist_str = get_eval_field_distribution(st.session_state["df"])
                            if st.session_state.get("template_type") == "GSB对比":
                                prompt = PARSE_METRIC_GSB_PROMPT.format(field_distribution=field_dist_str, metric_description=spec_content)
                            else:  # "线上评测" 和 "等级对比" 均使用标准 Excel 解析
                                prompt = PARSE_METRIC_EXCEL_PROMPT.format(field_distribution=field_dist_str, excel_content=spec_content)
                            response = llm.call_text(prompt, max_tokens=8192)
                            parsed_result = llm.parse_json_response(response)
                            parsed_result = auto_fix_operators(parsed_result, st.session_state["df"])
                            parsed_result, fix_log = auto_fix_values(parsed_result, st.session_state["df"])
                            parsed_result, level_fix_log = auto_fix_level_codes(parsed_result, st.session_state["df"], spec_content)
                            fix_log = fix_log + level_fix_log
                            parsed_result = migrate_or_conditions_to_flat(parsed_result)
                            st.session_state["parsed_metrics"] = parsed_result
                            editing_copy = parsed_result.copy()
                            # 保存 AI 分析出的分母条件，供"AI分析的分母"选项使用
                            ai_denom = parsed_result.get("common_denominator", {})
                            ai_denom_conds = ai_denom.get("conditions", [])
                            if ai_denom_conds:
                                editing_copy.setdefault("common_denominator", {})["ai_analyzed_conditions"] = ai_denom_conds
                                editing_copy["common_denominator"]["type"] = "ai"
                            else:
                                editing_copy.setdefault("common_denominator", {})["ai_analyzed_conditions"] = []
                                editing_copy["common_denominator"]["type"] = "all"
                            st.session_state["editing_metrics"] = editing_copy
                            st.success("✅ 解析成功！")
                            if fix_log:
                                fix_lines = []
                                for item in fix_log:
                                    source = item.get("source", "")
                                    suffix = f"（{source}）" if source else f"（相似度 {item['ratio']}%）"
                                    fix_lines.append(f'- 字段「{item["field"]}」: "{item["old"]}" → "{item["new"]}" {suffix}')
                                st.warning("⚠️ 以下值已自动修正：\n" + "\n".join(fix_lines))
                        except Exception as e:
                            import traceback
                            st.error(f"解析失败：{type(e).__name__}: {e}")
                            st.code(traceback.format_exc())
            except Exception as e:
                st.error(f"文件读取失败: {str(e)}")

        # 分组维度确认UI（始终显示，支持手动新增和删除）
        st.markdown("---")
        st.markdown("📊 **分组维度配置：**")
        if "auto_group_dimensions" not in st.session_state:
            st.session_state["auto_group_dimensions"] = []
        auto_groups = st.session_state["auto_group_dimensions"]

        confirmed_groups = []
        dims_to_remove = []
        for gi, gd in enumerate(auto_groups):
            col_label, col_field, col_vals, col_del = st.columns([2, 3, 4, 1])
            with col_label:
                new_label = st.text_input("维度名", value=gd.get("label", ""), key=f"group_dim_label_{gi}", label_visibility="collapsed")
            with col_field:
                if gd.get("field"):
                    field_idx = st.session_state["columns"].index(gd["field"]) if gd["field"] in st.session_state["columns"] else 0
                    selected_field = st.selectbox(
                        "匹配字段",
                        options=st.session_state["columns"],
                        index=field_idx,
                        key=f"group_dim_field_{gi}",
                        label_visibility="collapsed"
                    )
                    st.caption("✅ 自动匹配")
                else:
                    selected_field = st.selectbox(
                        "选择匹配字段",
                        options=["（请选择）"] + st.session_state["columns"],
                        index=0,
                        key=f"group_dim_field_{gi}",
                        label_visibility="collapsed"
                    )
                    st.caption("⚠️ 需手动选择")
                    if selected_field == "（请选择）":
                        selected_field = None
            with col_vals:
                available_vals = get_column_unique_values(
                    st.session_state["df"], selected_field
                ) if selected_field else []
                all_options = sorted(set(available_vals + gd["values"]))
                selected_vals = st.multiselect(
                    "子项", options=all_options,
                    default=[v for v in gd["values"] if v in all_options],
                    key=f"group_dim_vals_{gi}",
                    label_visibility="collapsed"
                )
            with col_del:
                if st.button("🗑️", key=f"del_group_dim_{gi}", help="删除此分组维度"):
                    dims_to_remove.append(gi)
            if gi not in dims_to_remove:
                confirmed_groups.append({"label": new_label or gd["label"], "field": selected_field, "values": selected_vals})

        # 删除操作
        if dims_to_remove:
            for idx in sorted(dims_to_remove, reverse=True):
                auto_groups.pop(idx)
            st.session_state["auto_group_dimensions"] = auto_groups
            st.rerun()

        # 手动新增按钮
        if st.session_state.get("columns"):
            if st.button("➕ 手动添加分组维度", key="add_group_dim"):
                st.session_state["_adding_group_dim"] = True
                st.rerun()

        if st.session_state.get("_adding_group_dim"):
            with st.container(border=True):
                st.markdown("**新增分组维度**")
                new_dim_col1, new_dim_col2 = st.columns([2, 3])
                with new_dim_col1:
                    new_dim_label = st.text_input("维度标签（如：坐资）", key="new_dim_label_input")
                with new_dim_col2:
                    new_dim_field = st.selectbox("对应数据字段", options=["（请选择）"] + st.session_state["columns"], key="new_dim_field_input")
                btn_col1, btn_col2 = st.columns(2)
                with btn_col1:
                    if st.button("✅ 确认添加", key="confirm_add_dim"):
                        if new_dim_label and new_dim_field and new_dim_field != "（请选择）":
                            vals = get_column_unique_values(st.session_state["df"], new_dim_field)
                            auto_groups.append({"label": new_dim_label, "field": new_dim_field, "values": vals})
                            st.session_state["auto_group_dimensions"] = auto_groups
                            st.session_state["_adding_group_dim"] = False
                            st.rerun()
                        else:
                            st.warning("请填写维度标签并选择字段")
                with btn_col2:
                    if st.button("取消", key="cancel_add_dim"):
                        st.session_state["_adding_group_dim"] = False
                        st.rerun()

        st.session_state["confirmed_group_dimensions"] = [g for g in confirmed_groups if g.get("field")]

        # ==================== 字段映射确认（人工介入点2） ====================
        if st.session_state.get("parsed_metrics"):
            render_field_mappings(
                st.session_state["parsed_metrics"],
                st.session_state["columns"]
            )

        # ==================== 解析结果确认区域 ====================
        st.divider()

        if st.session_state.get("parsed_metrics"):
            st.subheader("6. 确认口径配置")

            editing = st.session_state["editing_metrics"]

            # 口径编辑区域（fragment化，内部rerun不会滚动页面）
            render_metrics_editor_fragment()

            st.divider()

            col1, col2, col3 = st.columns([2, 1, 1])

            with col1:
                if st.button("确认口径配置", type="primary", use_container_width=True):
                    is_valid, errors = validate_metrics_config(editing, st.session_state["columns"])
                    if is_valid:
                        st.session_state["confirmed_metrics"] = editing.copy()
                        presets_bad, presets_good = derive_case_filters_from_metrics(editing)
                        st.session_state["badcase_presets"] = presets_bad
                        st.session_state["goodcase_presets"] = presets_good
                        st.success("口径配置已确认！")
                    else:
                        for error in errors:
                            st.error(error)

            with col2:
                if st.button("重置", use_container_width=True):
                    st.session_state["parsed_metrics"] = None
                    st.session_state["editing_metrics"] = EMPTY_METRICS_CONFIG.copy()
                    st.session_state.pop("confirmed_metrics", None)
                    st.session_state.pop("stats_result", None)
                    st.session_state.pop("generated_report", None)
                    st.session_state["is_generating_report"] = False
                    st.session_state["trigger_parse_excel"] = False
                    clear_metrics_editor_keys()
                    st.rerun()

            with col3:
                if st.button("查看配置详情", use_container_width=True):
                    show_config_dialog(editing)

            # ==================== 统计区域 ====================
            if st.session_state.get("confirmed_metrics"):
                st.divider()
                st.subheader("7. 执行统计")

                confirmed = st.session_state["confirmed_metrics"]

                comparison_mode = st.toggle("启用版本对比模式", value=st.session_state.get("comparison_mode", False), key="comparison_mode_toggle")
                st.session_state["comparison_mode"] = comparison_mode

                has_auto_groups = bool(st.session_state.get("confirmed_group_dimensions"))
                default_grouping = st.session_state.get("grouping_mode", has_auto_groups)
                grouping_mode = st.toggle("启用分组统计模式", value=default_grouping, key="grouping_mode_toggle")
                st.session_state["grouping_mode"] = grouping_mode

                if comparison_mode:
                    st.markdown("**对比配置**")
                    if not st.session_state["columns"]:
                        st.warning("当前数据没有可用于对比的字段，请先检查上传数据。")
                        compare_field = None
                        version_a = None
                        version_b = None
                    else:
                        col1, col2, col3 = st.columns(3)

                        with col1:
                            compare_field = st.selectbox("对比字段", options=st.session_state["columns"], key="compare_field_select")

                        compare_values = get_column_unique_values(st.session_state["df"], compare_field) if compare_field else []
                        if not compare_values:
                            st.warning("所选对比字段暂无可用值，请更换字段。")
                            version_a = None
                            version_b = None
                            with col2:
                                st.selectbox("版本A（基准）", options=["（暂无可用值）"], index=0, disabled=True, key="version_a_select_empty")
                            with col3:
                                st.selectbox("版本B（对比）", options=["（暂无可用值）"], index=0, disabled=True, key="version_b_select_empty")
                        else:
                            with col2:
                                version_a = st.selectbox("版本A（基准）", options=compare_values, key="version_a_select")

                            with col3:
                                version_b = st.selectbox("版本B（对比）", options=compare_values, index=1 if len(compare_values) > 1 else 0, key="version_b_select")
                else:
                    compare_field = None
                    version_a = None
                    version_b = None

                if grouping_mode:
                    st.markdown("**分组配置**")
                    auto_groups = st.session_state.get("confirmed_group_dimensions", [])
                    if auto_groups:
                        st.info(f"已从口径Excel自动识别 {len(auto_groups)} 个分组维度: {', '.join(g['label'] for g in auto_groups)}")
                        group_field = auto_groups[0]["field"]
                    else:
                        group_field = st.selectbox("分组字段", options=st.session_state["columns"], key="group_field_select")
                else:
                    group_field = None

                if st.button("开始统计", type="primary", use_container_width=True):
                    with st.spinner("正在计算统计指标..."):
                        try:
                            confirmed_obj = MetricsConfig.from_dict(confirmed) if isinstance(confirmed, dict) else confirmed
                            if comparison_mode and compare_field and version_a and version_b:
                                stats_result = DataService(st.session_state["df"]).calculate_metrics_with_comparison(confirmed_obj, compare_field, version_a, version_b)
                            elif grouping_mode:
                                auto_groups = st.session_state.get("confirmed_group_dimensions", [])
                                if auto_groups and len(auto_groups) > 1:
                                    multi_results = []
                                    for gd in auto_groups:
                                        r = DataService(st.session_state["df"]).calculate_metrics_with_grouping(confirmed_obj, gd["field"])
                                        multi_results.append(r)
                                    stats_result = {"multi_group": True, "group_results": multi_results}
                                elif group_field:
                                    stats_result = DataService(st.session_state["df"]).calculate_metrics_with_grouping(confirmed_obj, group_field)
                                else:
                                    stats_result = DataService(st.session_state["df"]).calculate_metrics(confirmed_obj)
                            else:
                                stats_result = DataService(st.session_state["df"]).calculate_metrics(confirmed_obj)

                            # 统一转换为 dict 供展示层使用
                            if hasattr(stats_result, 'to_dict') and stats_result.group_field:
                                # 分组模式：构建展示所需的 flat 格式
                                _groups = stats_result.groups
                                _gr_map = stats_result.group_results  # {group_value: StatsResult}

                                # 【关键修复】从配置中获取指标名称，避免因分组数据为空导致指标遗漏
                                _metric_names = [m.name for m in confirmed_obj.metrics] if hasattr(confirmed_obj, 'metrics') else []

                                # 如果配置中没有指标名称，从分组结果中收集（兼容旧逻辑）
                                if not _metric_names:
                                    for _g in _groups:
                                        _gs = _gr_map.get(_g)
                                        if _gs and hasattr(_gs, 'results'):
                                            for _r in _gs.results:
                                                if _r.name not in _metric_names:
                                                    _metric_names.append(_r.name)

                                # 构建 flat results（每个指标行包含各分组数据）
                                _flat_results = []
                                for _name in _metric_names:
                                    _row = {"name": _name}
                                    for _g in _groups:
                                        _gs = _gr_map.get(_g)
                                        if _gs and hasattr(_gs, 'results'):
                                            _mr = next((r for r in _gs.results if r.name == _name), None)
                                            _row[_g] = {"percentage": _mr.percentage, "numerator": _mr.numerator, "denominator": _mr.denominator} if _mr else {}
                                        else:
                                            _row[_g] = {}
                                    _flat_results.append(_row)
                                # 构建 denominator_counts
                                _denom_counts = {_g: _gr_map[_g].denominator_count for _g in _groups if _g in _gr_map}
                                stats_result = {
                                    "group_field": stats_result.group_field,
                                    "groups": _groups,
                                    "denominator_count": _denom_counts.get("全部", 0),
                                    "denominator_counts": _denom_counts,
                                    "denominator_description": stats_result.denominator_description,
                                    "results": _flat_results
                                }
                            elif hasattr(stats_result, 'to_dict'):
                                stats_result = stats_result.to_dict()
                            elif isinstance(stats_result, dict) and stats_result.get("multi_group"):
                                stats_result["group_results"] = [
                                    r.to_dict() if hasattr(r, 'to_dict') else r
                                    for r in stats_result["group_results"]
                                ]
                            st.session_state["stats_result"] = stats_result
                            st.success("统计完成！")
                        except Exception as e:
                            import traceback
                            st.error(f"统计失败: {str(e)}")
                            st.code(traceback.format_exc())

        else:
            st.info("👆 请先上传口径Excel")

        # ==================== 统计结果展示 ====================
        if st.session_state.get("stats_result"):
            st.divider()
            st.subheader("8. 统计结果")

            stats_result = st.session_state["stats_result"]
            is_comparison = st.session_state.get("comparison_mode", False)
            is_grouping = st.session_state.get("grouping_mode", False) and ("groups" in stats_result or stats_result.get("multi_group"))

            if is_grouping:
                spec_descs = st.session_state.get("spec_descriptions", {})
                if stats_result.get("multi_group"):
                    for idx, gr in enumerate(stats_result["group_results"]):
                        st.markdown(f"**分组字段: {gr.get('group_field', '')}**")
                        st.caption(f"分母描述: {gr.get('denominator_description', '未设置')}")
                        if spec_descs:
                            grouping_df = build_stats_table_from_spec(gr, spec_descs, is_grouping=True, grouping_result=gr)
                        else:
                            grouping_df = format_stats_result_for_grouping(gr)
                        st.dataframe(grouping_df, use_container_width=True, hide_index=True)
                        # 下载按钮
                        excel_path = ExportService().export_indicators_to_excel(grouping_df, f"分组统计_{gr.get('group_field', idx)}", f"temp/exported/分组统计_{idx}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                        with open(excel_path, "rb") as f:
                            st.download_button(
                                label=f"📥 导出此分组表Excel",
                                data=f,
                                file_name=f"分组统计_{gr.get('group_field', idx)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"download_grouping_{idx}"
                            )
                        st.markdown("")
                else:
                    st.markdown(f"**分组字段**: {stats_result.get('group_field', '')}  |  **分母描述**: {stats_result.get('denominator_description', '未设置')}")
                    # 显示各分组的分母数量（调试信息）
                    denom_counts = stats_result.get("denominator_counts", {})
                    if denom_counts:
                        count_str = " | ".join([f"{g}: {c}条" for g, c in denom_counts.items()])
                        st.caption(f"各分组分母数量: {count_str}")
                    if spec_descs:
                        grouping_df = build_stats_table_from_spec(stats_result, spec_descs, is_grouping=True, grouping_result=stats_result)
                    else:
                        grouping_df = format_stats_result_for_grouping(stats_result)
                    st.dataframe(grouping_df, use_container_width=True, hide_index=True)
                    # 下载按钮
                    excel_path = ExportService().export_indicators_to_excel(grouping_df, "分组统计结果", f"temp/exported/分组统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    with open(excel_path, "rb") as f:
                        st.download_button(
                            label="📥 导出分组统计表Excel",
                            data=f,
                            file_name=f"分组统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="download_grouping_single"
                        )
            elif is_comparison:
                st.markdown(f"**公共分母**: {stats_result.get('denominator_description', '未设置')}")
                col1, col2 = st.columns(2)
                with col1:
                    st.metric(f"版本A ({stats_result.get('version_a', '')}) 分母数", stats_result.get("denominator_count_a", 0))
                with col2:
                    st.metric(f"版本B ({stats_result.get('version_b', '')}) 分母数", stats_result.get("denominator_count_b", 0))
            else:
                st.metric("公共分母数", stats_result.get("denominator_count", 0))
                st.caption(f"分母描述: {stats_result.get('denominator_description', '未设置')}")

            if not is_grouping:
                st.markdown("**指标统计结果**")
                # 诊断信息：显示实际计算的指标
                actual_metrics = [r.get("name", "") for r in stats_result.get("results", [])]
                if actual_metrics:
                    with st.expander("📋 已计算的指标列表（点击展开）", expanded=False):
                        st.write(", ".join(actual_metrics))
                        st.caption(f"共 {len(actual_metrics)} 个指标")
                spec_descs = st.session_state.get("spec_descriptions", {})
                if spec_descs and not is_comparison:
                    results_df = build_stats_table_from_spec(stats_result, spec_descs)
                else:
                    results_df = format_stats_result_for_display(stats_result, is_comparison)
                    if spec_descs and "指标名称" in results_df.columns:
                        for col_name in reversed(["口径说明", "分项内容", "统计维度"]):
                            results_df.insert(0, col_name, results_df["指标名称"].map(
                                lambda n, cn=col_name: spec_descs.get(n, {}).get(cn, "")
                            ))

            if not is_grouping:
                if is_comparison:
                    style_subset = ['Gap(%)']
                    styled_df = results_df.style.applymap(color_gap, subset=style_subset)

                    def highlight_zero_denominator(row):
                        if row['版本A分母'] == 0 or row['版本B分母'] == 0:
                            return ['background-color: #f7efe2'] * len(row)
                        return [''] * len(row)

                    styled_df = styled_df.apply(highlight_zero_denominator, axis=1)
                    st.dataframe(styled_df, use_container_width=True, hide_index=True)
                    st.caption("趋势说明: ↑ 改善 | ↓ 退化 | → 持平")
                else:
                    styled_df = results_df.style.applymap(color_percentage, subset=['原始百分比'])

                    def highlight_zero_denominator(row):
                        if row['分母'] == 0:
                            return ['background-color: #f7efe2'] * len(row)
                        return [''] * len(row)

                    styled_df = styled_df.apply(highlight_zero_denominator, axis=1)
                    display_df = results_df.drop(columns=['原始百分比'], errors='ignore')
                    st.dataframe(display_df, use_container_width=True, hide_index=True)
                    st.caption("百分比颜色: 🟢 ≥90% | 🟠 60-90% | 🔴 <60%")

                # 下载按钮放在表格下方
                export_df = results_df.copy()
                if '原始百分比' in export_df.columns:
                    export_df = export_df.drop(columns=['原始百分比'])

                excel_path = ExportService().export_indicators_to_excel(export_df, "评测统计结果", f"temp/exported/统计结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                with open(excel_path, "rb") as f:
                    st.download_button(
                        label="📥 导出统计表Excel",
                        data=f,
                        file_name=f"统计结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_stats_non_grouping"
                    )

            # ==================== 生成评测报告 ====================
            st.divider()
            st.subheader("9. 生成评测报告")

            task_background = st.text_area(
                "任务背景",
                height=100,
                placeholder="例如：本次评测为自研图片生成模型3.1.0与3.1.1的对比评测，评测维度包括指令执行、图片生成质量、人物一致性等...",
                key="task_background_input"
            )

            col1, col2 = st.columns([1, 3])

            with col1:
                generate_report_clicked = st.button(
                    "生成评测报告",
                    type="primary",
                    use_container_width=True,
                    disabled=(not api_configured) or st.session_state.get("is_generating_report", False),
                    key="generate_report_btn"
                )
                if generate_report_clicked and api_configured and (not st.session_state.get("is_generating_report", False)):
                    st.session_state["is_generating_report"] = True

            with col2:
                if st.session_state.get("generated_report"):
                    st.info("报告已生成，可编辑或重新生成")

            if st.session_state.get("is_generating_report"):
                with st.spinner("AI正在生成评测报告..."):
                    try:
                        llm = LLMService(st.session_state["api_key"], st.session_state["base_url"], st.session_state.get("model_name", "glm-4"))

                        stats_text = format_stats_for_prompt(stats_result, is_comparison, is_grouping)

                        if is_comparison:
                            comparison_text = f"""
【版本对比信息】
- 对比字段: {stats_result.get('compare_field', '-')}
- 版本A: {stats_result.get('version_a', '-')}
- 版本B: {stats_result.get('version_b', '-')}
"""
                            comparison_section = "版本对比分析（对比两个版本的变化趋势，分析改善/退化的原因）"
                        else:
                            comparison_text = ""
                            comparison_section = "风险提示（指出需要关注的低指标）"

                        prompt = REPORT_PROMPT_TEMPLATE.format(
                            task_background=task_background or "未提供任务背景",
                            stats_text=stats_text,
                            comparison_text=comparison_text,
                            comparison_section=comparison_section
                        )

                        report = llm.call_text(prompt, temperature=0.7)
                        st.session_state["generated_report"] = report

                    except Exception as e:
                        st.error(format_api_error(e))
                st.session_state["is_generating_report"] = False

            # 报告编辑和导出
            if st.session_state.get("generated_report"):
                # 渲染预览
                st.markdown("**报告预览**")
                with st.container(border=True):
                    st.markdown(st.session_state["generated_report"])

                # 复制区域
                with st.expander("📋 复制报告内容", expanded=False):
                    st.code(st.session_state["generated_report"], language="markdown")
                    st.caption("点击右上角复制图标即可复制全文")

                # 编辑区域
                with st.expander("✏️ 编辑报告内容", expanded=False):
                    edited_report = st.text_area(
                        "报告内容",
                        value=st.session_state["generated_report"],
                        height=400,
                        key="report_editor",
                        label_visibility="collapsed"
                    )
                    st.session_state["generated_report"] = edited_report

                col1, col2, col3 = st.columns(3)

                with col1:
                    st.button("🔄 重新生成", use_container_width=True, on_click=clear_generated_report, key="regenerate_report_btn")

                with col2:
                    try:
                        word_path = ExportService().export_to_word(
                            st.session_state["generated_report"],
                            stats_result,
                            is_comparison,
                            "标注评测报告"
                        )

                        with open(word_path, "rb") as f:
                            st.download_button(
                                label="📄 导出为Word",
                                data=f,
                                file_name=f"评测报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx",
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                use_container_width=True
                            )
                    except Exception as e:
                        st.error(f"导出失败: {str(e)}")

            # ==================== Case筛选区域 ====================
            st.divider()
            st.subheader("10. Case筛选导出")

            case_tab1, case_tab2 = st.tabs(["❌ 问题案例", "✅ 优质案例"])

            with case_tab1:
                render_case_filter_tab(
                    st.session_state["df"],
                    "badcase",
                    "badcase_conditions",
                    "badcase_result"
                )

            with case_tab2:
                render_case_filter_tab(
                    st.session_state["df"],
                    "goodcase",
                    "goodcase_conditions",
                    "goodcase_result"
                )

    except Exception as e:
        st.error(f"文件读取失败: {str(e)}")
        with st.expander("查看错误详情"):
            st.exception(e)

else:
    st.info("👆 请先上传Excel文件开始使用")

    with st.expander("📖 使用说明"):
        st.markdown("""
        ### 工具使用流程

        1. **上传数据** - 上传标注好的Excel文件
        2. **配置API** - 在左侧配置模型的API信息
        3. **解析口径** - 输入或上传统计口径说明，AI自动解析
        4. **确认口径** - 人工检查并确认解析结果
        5. **执行统计** - 按口径自动计算各项指标
        6. **生成报告** - AI生成评测分析报告
        7. **Case筛选** - 筛选Bad Case / Good Case并导出

        ### Case筛选功能

        - **Bad Case**: 筛选问题案例，用于问题归因
        - **Good Case**: 筛选优秀案例，用于最佳实践总结
        - 支持多条件AND组合筛选
        - 支持导出完整筛选结果为Excel
        """)
